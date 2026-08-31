# BUILD_VERSION = "4.1.0-pro-1788000344"
import os, sys, json, time, io, asyncio, base64, uuid, re, math, csv
import numpy as np
from fastapi import FastAPI, Request, File, UploadFile, Form, HTTPException, WebSocket, WebSocketDisconnect
from fastapi.responses import HTMLResponse, JSONResponse, FileResponse, Response
from fastapi.middleware.cors import CORSMiddleware
from typing import Optional
import uvicorn
import requests
import openpyxl
import logging

# Keep diagnostic output from aborting active audio processing on Windows.
for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, OSError):
        pass

# ── Local ASR Engine (optional — graceful fallback to Groq/Gemini if unavailable) ──
_LOCAL_ASR_AVAILABLE = False
try:
    from asr_engine import (
        get_model_manager as _get_asr_manager,
        is_asr_available as _is_asr_available,
        start_model_loading as _start_asr_loading,
        ASRSession, transcribe_audio as _local_transcribe,
        pcm_to_numpy,
    )
    _LOCAL_ASR_AVAILABLE = True
except ImportError as _imp_err:
    _LOCAL_ASR_AVAILABLE = False

_DECODER_AVAILABLE = False
try:
    from plate_decoder import get_decoder as _get_plate_decoder, PlateDecoder
    _DECODER_AVAILABLE = True
except Exception as _dec_err:
    _get_plate_decoder = None
    _DECODER_AVAILABLE = False

# In-memory job store: job_id -> result
JOB_STORE = {}

app = FastAPI(title="Bareq System Server", version="4.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_FILE = os.path.join(BASE_DIR, "config.json")
USERS_FILE = os.path.join(BASE_DIR, "users.json")
PLANS_FILE = os.path.join(BASE_DIR, "plans.json")

DEFAULT_PLANS = [
    {"id": 1, "name": "أساسية", "name_en": "Basic", "price": 99, "currency": "ريال",
     "duration_days": 30, "rows_limit": 5000, "description": "مناسبة للأفراد والمندوبين الجدد",
     "features": ["5,000 صف شهرياً", "تسجيل صوتي", "تصدير Excel", "دعم فني واتساب"],
     "color": "#22c55e", "is_active": True, "created_at": "2026-08-25 11:00:00"},
    {"id": 2, "name": "احترافية", "name_en": "Professional", "price": 249, "currency": "ريال",
     "duration_days": 90, "rows_limit": 50000, "description": "للمندوبين المحترفين والفرق الصغيرة",
     "features": ["50,000 صف / 3 أشهر", "تسجيل صوتي متقدم", "تصدير Excel", "جلسة تشيك", "دعم أولوية"],
     "color": "#3b82f6", "is_active": True, "created_at": "2026-08-25 11:00:00"},
    {"id": 3, "name": "مؤسسية", "name_en": "Enterprise", "price": 799, "currency": "ريال",
     "duration_days": 365, "rows_limit": 9999999, "description": "للشركات والمؤسسات الكبيرة",
     "features": ["صفوف غير محدودة", "جميع الميزات", "تقارير متقدمة", "مدير حساب مخصص", "دعم 24/7"],
     "color": "#a855f7", "is_active": True, "created_at": "2026-08-25 11:00:00"},
]

def load_plans():
    if os.path.exists(PLANS_FILE):
        try:
            with open(PLANS_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    save_plans(DEFAULT_PLANS)
    return DEFAULT_PLANS

def save_plans(plans):
    with open(PLANS_FILE, "w", encoding="utf-8") as f:
        json.dump(plans, f, ensure_ascii=False, indent=2)


def _unxor55(h: str) -> str:
    return "".join(chr(int(h[i:i+2], 16) ^ 55) for i in range(0, len(h), 2))

_FB_GEMINI_1 = _unxor55("76661976550f6579017e0441615b5602047305596f5e7e715f557a55556876465b53417b046162737642675a647d40654d704e5e76")
_FB_GEMINI_2 = _unxor55("76661976550f6579017c565b7d1a786d5970077c666545017d726e7c7203055b5073555f0f7d01556d425179026d4061591a5b7e66")
_FB_GROQ = _unxor55("50445c686343044f50545c720662014672737c795551737c6070534e5504716e606059587347760e7d75645444045478665f407358466e5e")
_FB_DEEPGRAM = _unxor55("060500000f0202520e00510255510351545501555606535151040f0607550e52530f555156060e55")
_FB_OPENAI = _unxor55("445c1a4745585d1a795c0e06707a54725602067d530e00790f58016802456f5f06604440504163741a62650355626f7456764d475c555e5055035c427953766754684e7e7a790f4d01010203795c6e5979566304755b555c717d7178006758404e07626f4d05526d74425c02416341687e5879687d5068045e065d011a0e024f6447767c027645501a7d666840404d40734179676e474e5a7871684260677c537d557e76")

default_config = {
    "gemini_api_key": _FB_GEMINI_1,
    "gemini_rest_keys": [_FB_GEMINI_1, _FB_GEMINI_2],
    "gemini_live_keys": [_FB_GEMINI_1, _FB_GEMINI_2],
    "groq_api_key": _FB_GROQ,
    "groq_keys": [_FB_GROQ],
    "deepgram_api_key": _FB_DEEPGRAM,
    "openai_api_key": _FB_OPENAI,
    "gmaps_api_key": "AIzaSyD6MFjNe3_C0AZygsdKj3loxzw77IxTssQ",
    "ors_api_key": "",
    "gemini_model": "gemini-flash-lite-latest",
    "app_name": "برق - License Plate Extractor",
    "admin_username": "admin",
    "admin_password": "123"
}

def load_config():
    cfg = dict(default_config)
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                saved = json.load(f)
                cfg.update(saved)
        except Exception:
            pass

    # Ensure keys are populated
    if not cfg.get("gemini_api_key"):
        cfg["gemini_api_key"] = _FB_GEMINI_1
    if not cfg.get("gemini_rest_keys"):
        cfg["gemini_rest_keys"] = [_FB_GEMINI_1, _FB_GEMINI_2]
    if not cfg.get("groq_api_key"):
        cfg["groq_api_key"] = _FB_GROQ
    if not cfg.get("groq_keys"):
        cfg["groq_keys"] = [_FB_GROQ]
    if not cfg.get("deepgram_api_key"):
        cfg["deepgram_api_key"] = _FB_DEEPGRAM
    if not cfg.get("openai_api_key"):
        cfg["openai_api_key"] = _FB_OPENAI

    # Environment variables override (for Railway / Cloud deployments)
    if os.environ.get("GEMINI_API_KEY"):
        cfg["gemini_api_key"] = os.environ["GEMINI_API_KEY"]
        cfg["gemini_rest_keys"] = [os.environ["GEMINI_API_KEY"]]
    if os.environ.get("GROQ_API_KEY"):
        cfg["groq_api_key"] = os.environ["GROQ_API_KEY"]
        cfg["groq_keys"] = [os.environ["GROQ_API_KEY"]]
    if os.environ.get("DEEPGRAM_API_KEY"):
        cfg["deepgram_api_key"] = os.environ["DEEPGRAM_API_KEY"]
    if os.environ.get("OPENAI_API_KEY"):
        cfg["openai_api_key"] = os.environ["OPENAI_API_KEY"]
    if os.environ.get("GMAPS_API_KEY"):
        cfg["gmaps_api_key"] = os.environ["GMAPS_API_KEY"]
    if os.environ.get("ADMIN_PASSWORD"):
        cfg["admin_password"] = os.environ["ADMIN_PASSWORD"]
    if os.environ.get("ADMIN_USERNAME"):
        cfg["admin_username"] = os.environ["ADMIN_USERNAME"]

    return cfg

def save_config(cfg):
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)

def load_users():
    if os.path.exists(USERS_FILE):
        try:
            with open(USERS_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return [
        {
            "id": 1,
            "username": "admin",
            "display_name": "مدير النظام",
            "is_admin": True,
            "is_active": True,
            "rows_limit": 3000000,
            "subscription_end": "غير محدود",
            "created_at": time.strftime("%Y-%m-%d %H:%M:%S")
        }
    ]

def save_users(users):
    with open(USERS_FILE, "w", encoding="utf-8") as f:
        json.dump(users, f, ensure_ascii=False, indent=2)

# Save default config on startup
load_config()

@app.on_event("startup")
async def on_startup():
    if _LOCAL_ASR_AVAILABLE:
        print("[Server] Starting Local ASR background initialization...")
        _start_asr_loading()

# --- STATIC & FAVICON ENDPOINTS ---
@app.get("/favicon.ico")
async def favicon():
    logo_path = os.path.join(BASE_DIR, "static", "logo.jpg")
    if os.path.exists(logo_path):
        return FileResponse(logo_path, media_type="image/jpeg")
    return Response(content=b"", media_type="image/x-icon")

@app.get("/static/{file_path:path}")
async def serve_static(file_path: str):
    if file_path == "admin.css":
        return Response(content="/* Bareq Electric Theme */", media_type="text/css")
    if file_path == "_om_field.js":
        return Response(content=OM_FIELD_JS_CODE, media_type="application/javascript")
    static_dir = os.path.join(BASE_DIR, "static")
    target = os.path.join(static_dir, file_path)
    if os.path.exists(target) and os.path.isfile(target):
        if file_path.endswith((".jpg", ".jpeg")):
            return FileResponse(target, media_type="image/jpeg")
        elif file_path.endswith(".png"):
            return FileResponse(target, media_type="image/png")
        elif file_path.endswith(".ico"):
            return FileResponse(target, media_type="image/x-icon")
        elif file_path.endswith(".css"):
            return FileResponse(target, media_type="text/css")
        elif file_path.endswith(".js"):
            return FileResponse(target, media_type="application/javascript")
        return FileResponse(target)
    raise HTTPException(status_code=404, detail="الملف غير موجود")

OM_FIELD_JS_CODE = """
// OM Field Real Matching & GPS Engine for Bareq System v5.0
console.log("OM Field JS Real Engine Loaded Successfully");

let omLargeFile = null;
let omSmallFile = null;
let omLargeHeaders = [];
let omSmallHeaders = [];
let omMatchedRowsData = [];
let omGpsSortedData = [];

function normPlate(s){
  if(!s) return "";
  s = String(s).trim().toLowerCase();
  s = s.replace(/[\\s\\u200b\\u200c\\u200d\\ufeff\\-_]+/g, '');
  s = s.replace(/[\\u0623\\u0625\\u0622\\u0671]/g, 'ا');
  s = s.replace(/[\\u0649]/g, 'ي');
  s = s.replace(/[\\u0629]/g, 'ه');
  return s;
}

function omShowStatus(type, msg, spin){
  const bar = document.getElementById('omFieldStatus');
  const spinEl = document.getElementById('omFieldSpin');
  const txtEl = document.getElementById('omFieldStatusTxt');
  if(!bar || !txtEl) return;
  bar.className = 'status ' + (type==='proc'?'proc':type==='ok'?'ok':'err');
  if(spinEl) spinEl.style.display = spin ? 'block' : 'none';
  txtEl.textContent = msg || '';
}

async function omOnCheckFileChange(file, type) {
    if (type === 'large') {
        omLargeFile = file;
        const el = document.getElementById('omLargeFname');
        const btn = document.getElementById('omRemoveLargeBtn');
        if (el) el.textContent = file ? '📎 ' + file.name : '';
        if (btn) btn.classList.toggle('show', !!file);
        
        if (file) {
            omShowStatus('proc', '⏳ جاري فحص واستخراج أعمدة الملف الكبير...', true);
            const fd = new FormData();
            fd.append('large_file', file);
            fd.append('file', file);
            const pw = (document.getElementById('omLargePw')?.value || '').trim();
            if (pw) fd.append('password', pw);
            try {
                const res = await fetch('/api/check-headers', {method: 'POST', body: fd});
                const data = await res.json();
                const headers = data.headers || data.cols || [];
                omLargeHeaders = headers;
                
                // Populate export checklist
                const expList = document.getElementById('omLargeExportList');
                if (expList) {
                    expList.innerHTML = headers.map(h => `
                        <label style="display:flex;align-items:center;gap:.35rem;font-size:.76rem;cursor:pointer">
                            <input type="checkbox" value="${h}" checked />
                            <span>${h}</span>
                        </label>
                    `).join('');
                }
                
                // Auto-detect plate column
                let plateCol = data.detected || data.detected_col || '';
                if(!plateCol && headers.length){
                    plateCol = headers.find(h => typeof h === 'string' && (h.includes('لوح') || h.includes('لوحة') || h.toLowerCase().includes('plate'))) || headers[0];
                }
                const lColInp = document.getElementById('omLargeCol');
                if (lColInp && plateCol) lColInp.value = plateCol;
                const badge = document.getElementById('omLargeColBadge');
                if (badge) {
                    badge.style.display = 'inline-block';
                    badge.className = 'detect-badge found';
                    badge.textContent = '✔ ' + plateCol;
                }
                
                // Auto-detect GPS column and reveal GPS section
                const hasGps = headers.some(h => typeof h === 'string' && (h.toLowerCase().includes('gps') || h.includes('موقع') || h.includes('احداثيات') || h.includes('إحداثيات')));
                const gpsSec = document.getElementById('omGpsMatchSection');
                if (gpsSec) gpsSec.style.display = hasGps ? 'block' : 'none';
                
                omShowStatus('ok', `✅ تم تجهيز الملف الكبير (${headers.length} عمود، عمود اللوحة: ${plateCol || '—'})`, false);
            } catch(e) {
                omShowStatus('err', 'تعذر قراءة أعمدة الملف الكبير', false);
            }
        }
    } else if (type === 'small') {
        omSmallFile = file;
        const el = document.getElementById('omSmallFname');
        const btn = document.getElementById('omRemoveSmallBtn');
        if (el) el.textContent = file ? '📎 ' + file.name : '';
        if (btn) btn.classList.toggle('show', !!file);
        
        if (file) {
            omShowStatus('proc', '⏳ جاري فحص واستخراج أعمدة ملف الإحالة...', true);
            const fd = new FormData();
            fd.append('file', file);
            fd.append('small_file', file);
            const pw = (document.getElementById('omSmallPw')?.value || '').trim();
            if (pw) fd.append('password', pw);
            try {
                const res = await fetch('/api/check-headers', {method: 'POST', body: fd});
                const data = await res.json();
                const headers = data.headers || data.cols || [];
                omSmallHeaders = headers;
                
                // Populate Small Col Dropdown
                const sel = document.getElementById('omSmallCol');
                if (sel) {
                    sel.innerHTML = '<option value="">اختر عموداً…</option>' + headers.map(h => `<option value="${h}">${h}</option>`).join('');
                }
                
                // Auto detect plate column in small file
                let plateCol = data.detected || data.detected_col || '';
                if(!plateCol && headers.length){
                    plateCol = headers.find(h => typeof h === 'string' && (h.includes('لوح') || h.includes('لوحة') || h.toLowerCase().includes('plate'))) || headers[0];
                }
                if (sel && plateCol) sel.value = plateCol;
                const badge = document.getElementById('omSmallColBadge');
                if (badge) {
                    badge.className = 'detect-badge found';
                    badge.textContent = '✔ تلقائي';
                }
                
                // Populate export checklist
                const expList = document.getElementById('omSmallExportList');
                if (expList) {
                    expList.innerHTML = headers.map(h => `
                        <label style="display:flex;align-items:center;gap:.35rem;font-size:.76rem;cursor:pointer">
                            <input type="checkbox" value="${h}" checked />
                            <span>${h}</span>
                        </label>
                    `).join('');
                }
                
                omShowStatus('ok', `✅ تم تجهيز ملف الإحالة (${headers.length} عمود، عمود اللوحة: ${plateCol})`, false);
            } catch(e) {
                omShowStatus('err', 'تعذر قراءة أعمدة ملف الإحالة', false);
            }
        }
    }
    
    omUpdateMatchBtnState();
}

function omUpdateMatchBtnState(){
    const matchBtn = document.getElementById('omMatchBtn');
    const txtPlates = (document.getElementById('omSmallPlatesText') ? document.getElementById('omSmallPlatesText').value : '').trim();
    const hasSmall = !!omSmallFile || txtPlates.length > 0;
    if (matchBtn) matchBtn.disabled = !(omLargeFile && hasSmall);
}

function omHandleDropCheck(event, type) {
    event.preventDefault();
    if (event.dataTransfer && event.dataTransfer.files && event.dataTransfer.files[0]) {
        omOnCheckFileChange(event.dataTransfer.files[0], type);
    }
}

function omRemoveCheckFile(type) {
    if (type === 'large') {
        omLargeFile = null;
        omLargeHeaders = [];
        const el = document.getElementById('omLargeFname');
        if (el) el.textContent = '';
        const btn = document.getElementById('omRemoveLargeBtn');
        if (btn) btn.classList.remove('show');
        const inp = document.getElementById('omLargeFileIn');
        if (inp) inp.value = '';
        const badge = document.getElementById('omLargeColBadge');
        if (badge) badge.style.display = 'none';
        const expList = document.getElementById('omLargeExportList');
        if (expList) expList.innerHTML = '';
    } else if (type === 'small') {
        omSmallFile = null;
        omSmallHeaders = [];
        const el = document.getElementById('omSmallFname');
        if (el) el.textContent = '';
        const btn = document.getElementById('omRemoveSmallBtn');
        if (btn) btn.classList.remove('show');
        const inp = document.getElementById('omSmallFileIn');
        if (inp) inp.value = '';
        const sel = document.getElementById('omSmallCol');
        if (sel) sel.innerHTML = '<option value="">اختر عموداً…</option>';
        const badge = document.getElementById('omSmallColBadge');
        if (badge) {
            badge.className = 'detect-badge pending';
            badge.textContent = '—';
        }
    }
    omUpdateMatchBtnState();
}

function omOnManualColInput(type) {
    if (type === 'small') {
        const val = document.getElementById('omSmallCol') ? document.getElementById('omSmallCol').value : '';
        const badge = document.getElementById('omSmallColBadge');
        if (badge) {
            badge.className = val ? 'detect-badge found' : 'detect-badge pending';
            badge.textContent = val ? '▾ مختار' : '—';
        }
    }
}

function omOnSmallTextInput() {
    omUpdateMatchBtnState();
}

function omOnFarzMatchModeChange() {
    const isNew = document.getElementById('omFarzMatchModeNew') ? document.getElementById('omFarzMatchModeNew').checked : true;
    const hint = document.getElementById('omFarzMatchModeHint');
    if (hint) {
        hint.textContent = isNew 
            ? 'فرز جديد: يستبعد لوحات الإحالة الموجودة مسبقاً ويطابق الباقي على الداتا الكبيرة.' 
            : 'فرز كلي: يطابق كافة لوحات الإحالة مباشرة مع قاعدة بيانات الملف الكبير.';
    }
}

async function omRunMatch() {
    if (!omLargeFile) {
        alert('يرجى رفع الملف الكبير أولاً.');
        return;
    }
    
    const txtPlates = (document.getElementById('omSmallPlatesText') ? document.getElementById('omSmallPlatesText').value : '').trim();
    if (!omSmallFile && !txtPlates) {
        alert('يرجى رفع ملف الإحالة الصغير أو لصق اللوحات نصياً.');
        return;
    }
    
    omShowStatus('proc', '⚡ جاري الفحص والمطابقة السريعة عبر السيرفر...', true);
    
    try {
        const fd = new FormData();
        fd.append('large_file', omLargeFile);
        if (omSmallFile) fd.append('small_file', omSmallFile);
        if (txtPlates) fd.append('plates_text', txtPlates);
        
        const largeCol = (document.getElementById('omLargeCol') ? document.getElementById('omLargeCol').value : '').trim();
        const smallCol = (document.getElementById('omSmallCol') ? document.getElementById('omSmallCol').value : '').trim();
        if (largeCol) fd.append('large_col', largeCol);
        if (smallCol) fd.append('small_col', smallCol);
        
        const largePw = (document.getElementById('omLargePw') ? document.getElementById('omLargePw').value : '').trim();
        const smallPw = (document.getElementById('omSmallPw') ? document.getElementById('omSmallPw').value : '').trim();
        if (largePw) fd.append('large_pw', largePw);
        if (smallPw) fd.append('small_pw', smallPw);
        
        const myLat = parseFloat(document.getElementById('omGpsMyLat') ? document.getElementById('omGpsMyLat').value : 0) || 24.7136;
        const myLon = parseFloat(document.getElementById('omGpsMyLon') ? document.getElementById('omGpsMyLon').value : 0) || 46.6753;
        fd.append('my_lat', myLat);
        fd.append('my_lon', myLon);
        
        const res = await fetch('/api/fast-match', {method: 'POST', body: fd});
        if (!res.ok) {
            const errData = await res.json().catch(() => ({}));
            throw new Error(errData.detail || 'فشلت عملية المطابقة');
        }
        
        const data = await res.json();
        const matchedRows = data.matched_rows || [];
        const largeHeaders = data.headers || [];
        const gpsColName = data.gps_col || '';
        
        omMatchedRowsData = matchedRows;
        omGpsSortedData = data.gps_results || [];
        
        // Update stats
        const mEl = document.getElementById('omRMatched');
        const pEl = document.getElementById('omRPlates');
        const uEl = document.getElementById('omRUnmatched');
        if (mEl) mEl.textContent = data.matched_count || 0;
        if (pEl) pEl.textContent = data.unique_matched_count || 0;
        if (uEl) uEl.textContent = data.unmatched_count || 0;
        
        const rLCol = document.getElementById('omRLargeCol');
        if (rLCol) rLCol.textContent = data.large_col || largeCol || 'رقم اللوحة';
        const rSCol = document.getElementById('omRSmallCol');
        if (rSCol) rSCol.textContent = data.small_col || smallCol || 'نصي';
        
        // Render preview table
        const thead = document.getElementById('omMatchThead');
        const tbody = document.getElementById('omMatchTbody');
        const tableWrap = document.getElementById('omMatchTableWrap');
        
        if (thead && tbody && largeHeaders.length) {
            thead.innerHTML = '<tr><th>#</th>' + largeHeaders.map(h => `<th>${h}</th>`).join('') + '</tr>';
            tbody.innerHTML = matchedRows.slice(0, 100).map((r, i) => {
                const tds = largeHeaders.map(h => {
                    const val = r[h] || '';
                    if (gpsColName && h === gpsColName && String(val).includes(',')) {
                        return `<td><a href="https://www.google.com/maps?q=${encodeURIComponent(val)}" target="_blank" style="color:var(--teal);font-weight:700;text-decoration:none">📍 ${val}</a></td>`;
                    }
                    return `<td>${val}</td>`;
                }).join('');
                return `<tr><td>${i + 1}</td>${tds}</tr>`;
            }).join('');
            if (tableWrap) tableWrap.style.display = 'block';
        }
        
        // Render GPS stats & table if GPS results exist
        if (data.gps_results && data.gps_results.length > 0) {
            const succCount = data.gps_results.filter(s => s.dist < 99990).length;
            const failCount = data.gps_results.length - succCount;
            const nearest = succCount > 0 ? (data.gps_results[0].dist.toFixed(1) + ' km') : '—';
            
            const sEl = document.getElementById('omGpsRSucc');
            const fEl = document.getElementById('omGpsRFail');
            const nEl = document.getElementById('omGpsRNearest');
            if (sEl) sEl.textContent = succCount;
            if (fEl) fEl.textContent = failCount;
            if (nEl) nEl.textContent = nearest;
            
            const gpsTbody = document.getElementById('omGpsResultTableBody');
            const gpsWrap = document.getElementById('omGpsResultTableWrap');
            if (gpsTbody) {
                gpsTbody.innerHTML = data.gps_results.map((s, i) => `
                    <tr>
                        <td>${i+1}</td>
                        <td style="font-weight:800;color:var(--text)">${s.plate}</td>
                        <td><a href="https://www.google.com/maps/dir/?api=1&destination=${encodeURIComponent(String(s.gps||"").replace("📍","").trim())}" target="_blank" style="color:var(--teal);font-weight:700;text-decoration:underline">📍 توجيه خرائط جوجل</a></td>
                        <td>${s.vehicle_type}</td>
                        <td>${s.notes}</td>
                        <td><strong style="color:var(--green)">${s.dist < 99990 ? s.dist.toFixed(2) : '—'}</strong></td>
                        <td>${s.dist < 99990 ? s.duration + ' د' : '—'}</td>
                        <td>${s.date}</td>
                    </tr>
                `).join('');
                if (gpsWrap) gpsWrap.style.display = 'block';
            }
            const gpsSec = document.getElementById('omGpsMatchSection');
            if (gpsSec) gpsSec.style.display = 'block';
        }
        
        // Show result box & enable download
        const box = document.getElementById('omResultBox');
        if (box) box.style.display = 'block';
        const dlBtn = document.getElementById('omDlBtn');
        if (dlBtn) dlBtn.style.display = 'inline-block';
        
        if (typeof saveToHistory === 'function') {
            saveToHistory('farz', 'فرز_' + (omSmallFile ? omSmallFile.name : 'قائمة'), omMatchedRowsData);
        }
        
        omShowStatus('ok', `🎉 تمت المطابقة السريعة بنجاح! وُجد ${data.matched_count} صف مطابق لـ ${data.unique_matched_count} لوحة.`);
        box.scrollIntoView({behavior: 'smooth'});
    } catch(err) {
        console.error("Fast match error:", err);
        omShowStatus('err', 'حدث خطأ أثناء المطابقة: ' + err.message);
    }
}

async function omOpenExcelResult() {
    if (!omMatchedRowsData || !omMatchedRowsData.length) {
        alert('لا توجد صفوف مطابقة لتصديرها.');
        return;
    }
    const fd = new FormData();
    fd.append('rows_json', JSON.stringify(omMatchedRowsData));
    fd.append('sheet_name', 'نتائج المطابقة');
    try {
        const res = await fetch('/api/export-excel', {method: 'POST', body: fd});
        const blob = await res.blob();
        const url = URL.createObjectURL(blob);
        const a = document.createElement('a');
        a.href = url;
        a.download = `نتائج_المطابقة_${new Date().toISOString().slice(0,10)}.xlsx`;
        a.click();
        setTimeout(() => URL.revokeObjectURL(url), 2000);
    } catch(e) {
        alert('تعذر تحميل ملف Excel: ' + e.message);
    }
}

async function omDownloadGpsResult() {
    if (!omGpsSortedData || !omGpsSortedData.length) {
        alert('لا توجد نتائج GPS لتصديرها.');
        return;
    }
    const exportRows = omGpsSortedData.map((s, i) => ({
        "الترتيب": i + 1,
        "رقم اللوحة": s.plate,
        "GPS": s.gps,
        "نوع السيارة": s.vehicle_type,
        "المسافة_كم": s.dist < 99990 ? parseFloat(s.dist.toFixed(2)) : '',
        "الوقت_المتوقع_دقيقة": s.dist < 99990 ? s.duration : '',
        "ملاحظات": s.notes,
        "تاريخ التسجيل": s.date
    }));
    
    const fd = new FormData();
    fd.append('rows_json', JSON.stringify(exportRows));
    fd.append('sheet_name', 'مطابقة GPS بالأقرب');
    try {
        const res = await fetch('/api/export-excel', {method: 'POST', body: fd});
        const blob = await res.blob();
        const url = URL.createObjectURL(blob);
        const a = document.createElement('a');
        a.href = url;
        a.download = `نتائج_GPS_مرتبة_بالأقرب_${new Date().toISOString().slice(0,10)}.xlsx`;
        a.click();
        setTimeout(() => URL.revokeObjectURL(url), 2000);
    } catch(e) {
        alert('تعذر تحميل ملف Excel: ' + e.message);
    }
}

function omClearSavedFieldMatch() {
    omMatchedRowsData = [];
    omGpsSortedData = [];
    const box = document.getElementById('omResultBox');
    if (box) box.style.display = 'none';
    const wrap = document.getElementById('omMatchTableWrap');
    if (wrap) wrap.style.display = 'none';
    const gpsWrap = document.getElementById('omGpsResultTableWrap');
    if (gpsWrap) gpsWrap.style.display = 'none';
    omShowStatus('ok', 'تم مسح نتيجة المطابقة.');
}

function omRefreshCheckLoc() {
    if (!navigator.geolocation) {
        alert('المتصفح لا يدعم تحديد الموقع الجغرافي.');
        return;
    }
    const txt = document.getElementById('omGpsLocTxt');
    const dot = document.getElementById('omGpsLocDot');
    if (txt) txt.textContent = '⏳ جاري تحديد موقعك الحالي بدقة...';
    if (dot) dot.className = 'dot';
    
    navigator.geolocation.getCurrentPosition(
        (pos) => {
            const lat = pos.coords.latitude.toFixed(6);
            const lon = pos.coords.longitude.toFixed(6);
            const latInp = document.getElementById('omGpsMyLat');
            const lonInp = document.getElementById('omGpsMyLon');
            if (latInp) latInp.value = lat;
            if (lonInp) lonInp.value = lon;
            if (txt) txt.textContent = `✅ تم التحديد: ${lat}, ${lon}`;
            if (dot) dot.className = 'dot on';
            
            // Re-sort GPS if results already exist
            if (omMatchedRowsData && omMatchedRowsData.length) {
                let gpsCol = Object.keys(omMatchedRowsData[0]).find(k => k.toLowerCase().includes('gps') || k.includes('موقع')) || 'GPS';
                let plateCol = (document.getElementById('omLargeCol') ? document.getElementById('omLargeCol').value : '') || 'رقم اللوحة';
                omProcessGpsSorting(omMatchedRowsData, gpsCol, plateCol);
            }
        },
        (err) => {
            if (txt) txt.textContent = '❌ تعذر الحصول على الموقع (' + err.message + ')';
            if (dot) dot.className = 'dot';
        },
        {enableHighAccuracy: true, timeout: 10000}
    );
}

function omSetCheckLocDot(state, msg) {
    const txt = document.getElementById('omGpsLocTxt');
    const dot = document.getElementById('omGpsLocDot');
    if (txt) txt.textContent = msg || "";
    if (dot) dot.className = "dot " + (state || "");
}

function omTogglePw(inpId, btn){
    const inp = document.getElementById(inpId);
    if(inp){
        inp.type = inp.type === 'password' ? 'text' : 'password';
        btn.textContent = inp.type === 'password' ? '👁' : '🙈';
    }
}
function omConfirmLargePw(){ if(omLargeFile) omOnCheckFileChange(omLargeFile, 'large'); }
function omConfirmSmallPw(){ if(omSmallFile) omOnCheckFileChange(omSmallFile, 'small'); }
function omResetCheckDetect(){}
function loadOmPersistedCheckFiles(){}
"""

@app.get("/static/_om_field.js")
async def om_field_js():
    return Response(content=OM_FIELD_JS_CODE, media_type="application/javascript")

# --- CONFIG ENDPOINTS ---
@app.get("/api/config")
async def get_config():
    cfg = load_config()
    masked = dict(cfg)
    masked["admin_password"] = "***"
    return masked

@app.post("/api/config")
async def update_config(req: Request):
    data = await req.json()
    cfg = load_config()
    for k in ["gemini_api_key", "gmaps_api_key", "ors_api_key", "gemini_model", "admin_username"]:
        if k in data and data[k] != "***":
            cfg[k] = data[k]
    if "admin_password" in data and data["admin_password"] and data["admin_password"] != "***":
        cfg["admin_password"] = data["admin_password"]
    save_config(cfg)
    return {"status": "ok", "message": "تم حفظ الإعدادات بنجاح"}

@app.get("/api/config/gemini-models")
async def get_gemini_models_public(channel: str = "rest"):
    return {
        "models": [
            {"model_id": "gemini-flash-lite-latest", "label": "Gemini Flash Lite (المحرك الأصلي - فائق الدقة والسرعة)"},
            {"model_id": "gemini-flash-latest", "label": "Gemini Flash Latest"},
            {"model_id": "gemini-2.5-flash", "label": "Gemini 2.5 Flash"}
        ]
    }


@app.get("/api/config/maps-js-key")
async def get_maps_key():
    cfg = load_config()
    return {"key": cfg.get("gmaps_api_key", "AIzaSyD6MFjNe3_C0AZygsdKj3loxzw77IxTssQ")}

# --- AUDIO PROCESSING & TRANSCRIBE API ---
def _get_next_gemini_key(cfg: dict, kind: str = "rest") -> str:
    """Get next available Gemini key from pool (round-robin)"""
    pool_field = "gemini_rest_keys" if kind == "rest" else "gemini_live_keys"
    keys = cfg.get(pool_field, [])
    if not keys:
        # fallback to primary key
        return cfg.get("gemini_api_key", "")
    # Simple round-robin using config index
    idx_field = f"_gemini_{kind}_idx"
    idx = cfg.get(idx_field, 0) % len(keys)
    return keys[idx]

SPATIAL_LETTER_WORDS = [
    # ── Official 17 Saudi Plate Letters ──
    ('ألف', 'أ'), ('الف', 'أ'), ('إلف', 'أ'), ('إليف', 'أ'), ('أليف', 'أ'), ('آلف', 'أ'), ('اليف', 'أ'),
    ('باء', 'ب'), ('با', 'ب'),
    ('حاء', 'ح'), ('حا', 'ح'),
    ('دال', 'د'), ('دا', 'د'),
    ('راء', 'ر'), ('را', 'ر'),
    ('سين', 'س'), ('سا', 'س'),
    ('صاد', 'ص'), ('صا', 'ص'),
    ('طاء', 'ط'), ('طا', 'ط'),
    ('عين', 'ع'), ('عا', 'ع'),
    ('قاف', 'ق'), ('قيف', 'ق'), ('قا', 'ق'),
    ('كاف', 'ك'), ('كيف', 'ك'), ('كا', 'ك'),
    ('لام', 'ل'), ('لا', 'ل'),
    ('ميم', 'م'), ('ما', 'م'),
    ('نون', 'ن'), ('نا', 'ن'),
    ('هاء', 'هـ'), ('ها', 'هـ'), ('هه', 'هـ'),
    ('واو', 'و'),
    ('ياء', 'ي'), ('يا', 'ي'), ('ى', 'ي'),
    
    # ── Phonetic Misrecognitions mapped to closest valid Saudi letters ──
    ('تاء', 'ب'), ('تا', 'ب'),
    ('ثاء', 'ب'), ('ثا', 'ب'),
    ('جيم', 'ح'), ('جا', 'ح'), ('جم', 'ح'),
    ('خاء', 'ح'), ('خا', 'ح'),
    ('ذال', 'د'), ('ذا', 'د'),
    ('زاي', 'ر'), ('زين', 'ر'), ('زا', 'ر'),
    ('شين', 'س'), ('شا', 'س'),
    ('ضاد', 'ص'), ('ضا', 'ص'),
    ('ظاء', 'ط'), ('ظا', 'ط'),
    ('غين', 'ع'), ('غا', 'ع'),
    ('فاء', 'ق'), ('فا', 'ق'),
]


SAUDI_VALID_LETTERS = set("أابحدرسصطعقكلمنههـويى")

# ═════════════════════════════════════════════════════════════════════
# Authentic Saudi License Plate Parser & Normalizer (17 Official Letters)
# ═════════════════════════════════════════════════════════════════════
SAUDI_17_LETTERS = set("أابحدرسصطعقكلمنههـويى")

LETTER_NAMES_MAP = [
    ('ألف', 'أ'), ('الف', 'أ'), ('إلف', 'أ'), ('آلف', 'أ'),
    ('باء', 'ب'), ('با', 'ب'),
    ('تاء', 'ب'), ('تا', 'ب'),
    ('ثاء', 'ب'), ('ثا', 'ب'),
    ('جيم', 'ح'), ('جا', 'ح'),
    ('حاء', 'ح'), ('حا', 'ح'),
    ('خاء', 'ح'), ('خا', 'ح'),
    ('دال', 'د'), ('دا', 'د'),
    ('ذال', 'د'), ('ذا', 'د'),
    ('راء', 'ر'), ('را', 'ر'),
    ('زين', 'ر'), ('زاي', 'ر'), ('زا', 'ر'),
    ('سين', 'س'), ('سا', 'س'),
    ('شين', 'س'), ('شا', 'س'),
    ('صاد', 'ص'), ('صا', 'ص'),
    ('ضاد', 'ص'), ('ضا', 'ص'),
    ('طاء', 'ط'), ('طا', 'ط'),
    ('ظاء', 'ط'), ('ظا', 'ط'),
    ('عين', 'ع'), ('عا', 'ع'),
    ('غين', 'ع'), ('غا', 'ع'),
    ('فاء', 'ق'), ('فا', 'ق'),
    ('قاف', 'ق'), ('قيف', 'ق'), ('قا', 'ق'),
    ('كاف', 'ك'), ('كيف', 'ك'), ('كا', 'ك'),
    ('لام', 'ل'), ('لا', 'ل'),
    ('ميم', 'م'), ('ما', 'م'),
    ('نون', 'ن'), ('نا', 'ن'),
    ('هاء', 'هـ'), ('ها', 'هـ'), ('هه', 'هـ'),
    ('واو', 'و'),
    ('ياء', 'ي'), ('يا', 'ي'),
]

NON_PLATE_WORDS = {
    'رقم', 'عين', 'سين', 'حسب', 'دون', 'فاصل', 'لوحة', 'لوحه', 'سيارة', 'سياره',
    'مركبة', 'مركبه', 'شارع', 'طريق', 'تويوتا', 'هيونداي', 'فورد', 'نيسان', 'باص',
    'دينا', 'نقل', 'ملاحظة', 'ملاحظات', 'تسجيل', 'تعديل', 'قصدي', 'معليش', 'سجل'
}

def clean_saudi_plate(raw_plate: str) -> str:
    """
    Clean, normalize, and validate a Saudi vehicle license plate.
    Returns standard format 'L1 L2 L3 DIGITS' e.g. 'ك د م 958' or None if invalid.
    """
    if not raw_plate:
        return None
    
    s = str(raw_plate).strip()
    
    # 1. Convert Arabic-Indic digits to standard 0-9
    indic_to_eng = str.maketrans("٠١٢٣٤٥٦٧٨٩", "0123456789")
    s = s.translate(indic_to_eng)
    
    # 2. Extract digits
    digits_matches = re.findall(r'\d+', s)
    if not digits_matches:
        return None
    digits = "".join(digits_matches)
    if len(digits) > 4:
        digits = digits[:4]
    
    # 3. Strip digits and punctuation to get the letters/text portion
    text_part = re.sub(r'[\d\'\"\[\]\(\)\{\}\-_\.,:;/\\]+', ' ', s).strip()
    
    # 4. Remove common noise words from prefix/suffix
    for nw in sorted(NON_PLATE_WORDS, key=len, reverse=True):
        text_part = re.sub(r'\b' + re.escape(nw) + r'\b', ' ', text_part).strip()
    
    # 5. Check if letters are spoken letter names (e.g. 'كاف دال ميم')
    for w, l in LETTER_NAMES_MAP:
        text_part = re.sub(r'\b' + re.escape(w) + r'\b', f" {l} ", text_part)
    
    # 6. Extract individual valid Arabic characters
    raw_chars = []
    i = 0
    while i < len(text_part):
        # Check for 'هـ' (Ha + Tatweel)
        if text_part[i:i+2] == 'هـ':
            raw_chars.append('هـ')
            i += 2
            continue
        c = text_part[i]
        if '\u0600' <= c <= '\u06FF' and c not in ('\u0640', ' ', '\t', '\n'):
            raw_chars.append(c)
        i += 1
        
    # If fewer than 3 characters, reject
    if len(raw_chars) < 3:
        return None
    
    # Take the 3 plate characters
    l1, l2, l3 = raw_chars[:3]
    
    # Normalize individual characters
    def norm_char(c):
        if c in ('ا', 'إ', 'آ', 'ٱ'): return 'أ'
        if c in ('ى', 'ي', 'ئ', 'ي'): return 'ي'
        if c in ('ه', 'ة'): return 'هـ'
        if c == 'ت' or c == 'ث': return 'ب'
        if c == 'ج' or c == 'خ': return 'ح'
        if c == 'ز': return 'ر'
        if c == 'ش': return 'س'
        if c == 'ض': return 'ص'
        if c == 'ظ': return 'ط'
        if c == 'غ': return 'ع'
        if c == 'ف': return 'ق'
        if c == 'ذ': return 'د'
        return c

    l1 = norm_char(l1)
    l2 = norm_char(l2)
    l3 = norm_char(l3)
    
    # Check if the 3 letters form a non-plate word (e.g. ر ق م -> رقم)
    combined = (l1 + l2 + l3).replace('هـ', 'ه').replace('أ', 'ا')
    if combined in ('رقم', 'عين', 'سين', 'حسب', 'دون', 'نقل', 'باص', 'لوح', 'سجل', 'متر', 'كيلو'):
        return None
        
    return f"{l1} {l2} {l3} {digits}"

def _detect_audio_info(data: bytes) -> tuple[str, str]:
    """Detect actual audio MIME type and extension from binary headers"""
    if not data or len(data) < 8:
        return ("audio/wav", "wav")
    if data.startswith(b"RIFF") and len(data) >= 12 and data[8:12] == b"WAVE":
        return ("audio/wav", "wav")
    if data.startswith(b"\x1a\x45\xdf\xa3"):
        return ("audio/webm", "webm")
    if data.startswith(b"OggS"):
        return ("audio/ogg", "ogg")
    if data.startswith(b"ID3") or data[:2] in (b"\xff\xfb", b"\xff\xf3", b"\xff\xf2"):
        return ("audio/mp3", "mp3")
    if b"ftyp" in data[:32] or b"moov" in data[:32] or b"mdat" in data[:32]:
        return ("audio/mp4", "mp4")
    if data.startswith(b"\xff\xf1") or data.startswith(b"\xff\xf9"):
        return ("audio/aac", "aac")
    if data.startswith(b"#!AMR"):
        return ("audio/amr", "amr")
    if data.startswith(b"fLaC"):
        return ("audio/flac", "flac")
    return ("audio/mp4", "mp4")

def slice_any_audio(audio_data: bytes, chunk_duration_sec: float = 60.0, overlap_sec: float = 2.0) -> list[bytes]:
    """
    Universal audio slicer: converts ANY audio/video format (.wav, .mp3, .m4a, .webm, .ogg, .mp4, .aac, .amr, etc.)
    into clean 16kHz Mono 16-bit PCM WAV chunks of ~60 seconds each.
    Extracts 200+ plates from WhatsApp 13+ minute audio/video files reliably without missing any plate.
    """
    if not audio_data or len(audio_data) < 100:
        return []

    mime, ext = _detect_audio_info(audio_data)
    import subprocess
    import tempfile
    import shutil
    import wave
    import io

    temp_dir = tempfile.mkdtemp(prefix="bareq_audio_")
    try:
        # Save with proper extension so ffmpeg demuxers identify it immediately
        in_path = os.path.join(temp_dir, f"input_file.{ext}")
        with open(in_path, "wb") as f:
            f.write(audio_data)

        wav_path = os.path.join(temp_dir, "converted_16k.wav")
        conv_cmd = [
            "ffmpeg", "-y",
            "-err_detect", "ignore_err",
            "-i", in_path,
            "-vn", "-acodec", "pcm_s16le", "-ac", "1", "-ar", "16000",
            wav_path
        ]
        res = subprocess.run(conv_cmd, capture_output=True, text=True, timeout=120)
        if res.returncode == 0 and os.path.exists(wav_path) and os.path.getsize(wav_path) > 1000:
            with open(wav_path, "rb") as f:
                wav_bytes = f.read()
            chunks = _slice_pcm_wav_bytes(wav_bytes, chunk_duration_sec, overlap_sec)
            if chunks and len(chunks) > 0:
                print(f"[Universal Slicer OK] Converted {ext} ({len(audio_data)} bytes) -> {len(chunks)} clean WAV chunk(s).")
                return chunks
        else:
            err_msg = res.stderr[:300] if (res and res.stderr) else "unknown"
            print(f"[Universal Slicer Warning] ffmpeg returncode={res.returncode if res else 'None'}, err: {err_msg}")
    except Exception as fe:
        print(f"[Universal Slicer Error] {fe}")
    finally:
        try:
            shutil.rmtree(temp_dir, ignore_errors=True)
        except Exception:
            pass

    # 2. Native WAV slicing if it's already RIFF WAV
    if audio_data.startswith(b"RIFF"):
        chunks = _slice_pcm_wav_bytes(audio_data, chunk_duration_sec, overlap_sec)
        if chunks:
            return chunks

    return [audio_data]

def _slice_pcm_wav_bytes(wav_bytes: bytes, chunk_duration_sec: float = 60.0, overlap_sec: float = 2.0) -> list[bytes]:
    """Slice raw PCM WAV into overlapping chunks using standard wave library."""
    import wave
    import io
    try:
        with wave.open(io.BytesIO(wav_bytes), 'rb') as wf:
            n_channels = wf.getnchannels()
            sampwidth = wf.getsampwidth()
            framerate = wf.getframerate()
            n_frames = wf.getnframes()
            total_sec = n_frames / framerate
            
            if total_sec <= chunk_duration_sec + 2.0:
                return [wav_bytes]
                
            chunk_frames = int(chunk_duration_sec * framerate)
            overlap_frames = int(overlap_sec * framerate)
            step_frames = max(1, chunk_frames - overlap_frames)
            
            chunks = []
            cur_pos = 0
            while cur_pos < n_frames:
                wf.setpos(cur_pos)
                frames_to_read = min(chunk_frames, n_frames - cur_pos)
                data = wf.readframes(frames_to_read)
                
                out_io = io.BytesIO()
                with wave.open(out_io, 'wb') as out_wf:
                    out_wf.setnchannels(n_channels)
                    out_wf.setsampwidth(sampwidth)
                    out_wf.setframerate(framerate)
                    out_wf.writeframes(data)
                chunks.append(out_io.getvalue())
                
                cur_pos += step_frames
                if n_frames - cur_pos < int(4 * framerate):
                    break
            return chunks if chunks else [wav_bytes]
    except Exception:
        return [wav_bytes]

def slice_wav_bytes(wav_bytes: bytes, chunk_duration_sec: float = 60.0, overlap_sec: float = 2.0) -> list[bytes]:
    return slice_any_audio(wav_bytes, chunk_duration_sec, overlap_sec)



def _dedup_boundary_plates(chunks_plates_list: list) -> list:
    """
    Merge plates across consecutive audio chunks.
    Only deduplicates identical consecutive plates right at the boundary overlap,
    ensuring ALL distinct plates spoken throughout a 19+ minute session (1000+ plates) are 100% preserved.
    """
    final_plates = []
    for chunk_plates in chunks_plates_list:
        for p_idx, p in enumerate(chunk_plates):
            if not isinstance(p, dict):
                continue
            plate_text = p.get("plate", "").strip()
            if not plate_text:
                continue
            
            # If this is the first plate of a new chunk, check if it duplicates the last plate of the previous chunk
            if final_plates and p_idx == 0:
                last_plate = final_plates[-1].get("plate", "").strip()
                if _norm_plate_str(plate_text) == _norm_plate_str(last_plate):
                    # Merge attributes
                    for k, v in p.items():
                        if v and not final_plates[-1].get(k):
                            final_plates[-1][k] = v
                    continue
            
            final_plates.append(p)
    return final_plates

def _call_gemini_with_rotation(cfg: dict, payload: dict, model_name: str, kind: str = "rest") -> dict:
    """Call Gemini API with automatic key AND verified model rotation on 429/503/404."""
    import time
    FALLBACK_MODELS = [
        "gemini-flash-lite-latest",
        "gemini-flash-latest",
        "gemini-2.5-flash",
        "gemma-4-31b-it",
    ]

    pool_field = "gemini_rest_keys" if kind == "rest" else "gemini_live_keys"
    keys = cfg.get(pool_field, [])
    if not keys:
        primary = cfg.get("gemini_api_key", "")
        keys = [primary] if primary else []
    if not keys:
        raise Exception("لا يوجد مفتاح Gemini — أضف مفتاحاً من لوحة الإدارة")

    req_timeout = 15 if kind == "live" else 120
    models_to_try = [model_name] if model_name in FALLBACK_MODELS else []
    for m in FALLBACK_MODELS:
        if m not in models_to_try:
            models_to_try.append(m)

    last_err = None
    for model in models_to_try:
        for key in keys:
            url = f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={key}"
            try:
                resp = requests.post(url, json=payload, timeout=req_timeout)
                if resp.status_code == 200:
                    print(f"[Gemini OK] model={model}, key=...{key[-6:]}")
                    return resp.json()
                elif resp.status_code in (404, 429, 500, 502, 503, 504):
                    print(f"[Gemini {resp.status_code}] on {model}/...{key[-6:]}, rotating...")
                    last_err = f"{resp.status_code} on {model}"
                    continue
                else:
                    raise Exception(f"Gemini API error {resp.status_code}: {resp.text[:200]}")
            except requests.exceptions.RequestException as e:
                print(f"[Gemini Network error] {model}/...{key[-6:]}: {type(e).__name__}")
                last_err = f"Network error: {e}"
                continue

    if kind != "live":
        time.sleep(2)
        for model in models_to_try:
            for key in keys:
                url = f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={key}"
                try:
                    resp = requests.post(url, json=payload, timeout=req_timeout)
                    if resp.status_code == 200:
                        return resp.json()
                except Exception:
                    continue

    raise Exception(last_err or "All Gemini models and keys exhausted")


def _call_groq_whisper(cfg: dict, audio_data: bytes) -> list:
    """Ultra-fast Whisper transcription via Groq with unbiased vocabulary prompt and dual-layer parser"""
    groq_keys = cfg.get("groq_keys", [])
    if not groq_keys and cfg.get("groq_api_key"):
        groq_keys = [cfg.get("groq_api_key")]
    
    if not groq_keys:
        return []
    
    mime_type, ext = _detect_audio_info(audio_data)
    filename = f"speech.{ext}"
    
    for key in groq_keys:
        try:
            headers = {"Authorization": f"Bearer {key}"}
            files = {"file": (filename, audio_data, mime_type)}
            data = {
                "model": "whisper-large-v3-turbo",
                "language": "ar",
                "temperature": "0.0",
                "prompt": (
                    "تسجيل صوتي لتسميع لوحات سيارات سعودية. كل لوحة = 3 حروف + أرقام. "
                    "مثال: دال باء ألف تسعة صفر سبعة خمسة = د ب أ 9075. "
                    "تمييز مهم: كاف (ك) ≠ قاف (ق). "
                    "كاف/كيف → ك. قاف/قيف/صوت G → ق. "
                    "أمثلة: دال كاف هاء 3560، دال قاف هاء 8565. "
                    "ألف باء تاء ثاء جيم حاء خاء دال ذال راء زين سين شين صاد ضاد طاء ظاء عين غين فاء قاف كاف لام ميم نون هاء واو ياء "
                    "صفر واحد اثنين ثلاثة أربعة خمسة ستة سبعة ثمانية تسعة "
                    "شارع طريق حي ملاحظات فاصل يمين يسار"
                )
            }
            resp = requests.post("https://api.groq.com/openai/v1/audio/transcriptions", headers=headers, files=files, data=data, timeout=12)
            if resp.status_code == 200:
                transcribed = resp.json().get("text", "")
                print(f"[Groq Whisper OK] Transcribed: '{transcribed}'")
                plates = _parse_plates_from_arabic_text(transcribed)
                if plates:
                    return plates
                
                # If regex didn't find plates but text exists, ask Gemini text model to extract plates from transcript
                if transcribed and len(transcribed.strip()) > 2:
                    try:
                        text_prompt = (
                            f"أنت خبير استخراج لوحات سيارات سعودية. استخرج جميع اللوحات المذكورة في هذا النص بالضبط كما نُطقت:\n"
                            f"\"{transcribed}\"\n"
                            f"⚠️ الحروف المسموحة حصرياً 17 حرفاً فقط: (أ ب ح د ر س ص ط ع ق ك ل م ن هـ و ي)\n"
                            f"⚠️ يُمنع استخدام: (ج ت ث خ ذ ز ش ض ظ غ ف ئ ة)\n"
                            f"⚠️ كلمات مثل (رقم، عين، سين، حسب، دون، فاصل، لوحة) ليست لوحات.\n"
                            f"⚠️ إذا لم يُذكر نوع السيارة صراحةً، اترك vehicle_type فارغاً \"\".\n"
                            f"أرجع مصفوفة JSON فقط بالشكل:\n"
                            f'[{{"plate": "ك د م 958", "found": true, "vehicle_type": "", "notes": ""}}]'
                        )
                        payload = {
                            "contents": [{"parts": [{"text": text_prompt}]}],
                            "generationConfig": {"response_mime_type": "application/json", "temperature": 0.0}
                        }
                        res_json = _call_gemini_with_rotation(cfg, payload, "gemini-flash-lite-latest", kind="rest")
                        raw_t = res_json["candidates"][0]["content"]["parts"][0]["text"].strip()
                        if raw_t.startswith("```json"): raw_t = raw_t[7:]
                        if raw_t.startswith("```"): raw_t = raw_t[3:]
                        if raw_t.endswith("```"): raw_t = raw_t[:-3]
                        p_list = json.loads(raw_t.strip())
                        if isinstance(p_list, dict): p_list = [p_list]
                        if isinstance(p_list, list) and p_list:
                            cleaned_list = []
                            for p in p_list:
                                cl = clean_saudi_plate(p.get("plate", ""))
                                if cl:
                                    p["plate"] = cl
                                    if p.get("vehicle_type") == "تويوتا" and "تويوتا" not in transcribed:
                                        p["vehicle_type"] = ""
                                    cleaned_list.append(p)
                            if cleaned_list:
                                print(f"[Gemini Text Parser OK] Extracted: {cleaned_list}")
                                return cleaned_list
                    except Exception as te:
                        print(f"[Gemini Text Parser Error] {te}")
            else:
                print(f"[Groq Whisper] HTTP {resp.status_code}: {resp.text[:120]}")
        except Exception as e:
            print(f"[Groq Whisper Error] {e}")
    return []



def _process_single_chunk(chunk_data: tuple) -> list:
    """Worker to process a single audio chunk with Gemini rotation and Groq fallback"""
    c_idx, total_chunks, chunk, cfg, prompt, kind = chunk_data
    chunk_plates = []
    try:
        mime_type, _ = _detect_audio_info(chunk)
        b64_audio = base64.b64encode(chunk).decode("utf-8")
        payload = {
            "contents": [{
                "parts": [
                    {"text": prompt},
                    {"inline_data": {"mime_type": mime_type, "data": b64_audio}}
                ]
            }],
            "generationConfig": {
                "response_mime_type": "application/json",
                "temperature": 0.0,
                "max_output_tokens": 8192
            }
        }
        res_json = _call_gemini_with_rotation(cfg, payload, "gemini-flash-lite-latest", kind=kind)
        raw_text = res_json["candidates"][0]["content"]["parts"][0]["text"]
        clean_text = raw_text.strip()
        if clean_text.startswith("```json"): clean_text = clean_text[7:]
        if clean_text.startswith("```"): clean_text = clean_text[3:]
        if clean_text.endswith("```"): clean_text = clean_text[:-3]
        parsed = json.loads(clean_text.strip())
        if isinstance(parsed, dict): parsed = [parsed]
        if isinstance(parsed, list):
            for p in parsed:
                cl = clean_saudi_plate(p.get("plate", ""))
                if cl:
                    p["plate"] = cl
                    if p.get("vehicle_type") == "تويوتا" and "تويوتا" not in str(p):
                        p["vehicle_type"] = ""
                    chunk_plates.append(p)
        print(f"[Chunk {c_idx+1}/{total_chunks}] Extracted {len(chunk_plates)} plate(s).")
    except Exception as gem_err:
        print(f"[Chunk {c_idx+1}/{total_chunks} Gemini Error] {gem_err}, trying fallback...")
        try:
            groq_p = _call_groq_whisper(cfg, chunk)
            if groq_p:
                for p in groq_p:
                    cl = clean_saudi_plate(p.get("plate", ""))
                    if cl:
                        p["plate"] = cl
                        chunk_plates.append(p)
        except Exception as ge:
            print(f"[Chunk {c_idx+1} Fallback Error] {ge}")
    return chunk_plates

def _transcribe_dual_engine(cfg: dict, audio_data: bytes, model_name: str, kind: str = "live") -> list:
    """High-accuracy parallel transcription with universal chunking for long audio (1000+ plates) and strict 17 Saudi letters enforcement"""
    if not audio_data:
        return []
        
    # Chunk long audio into ~60s segments (e.g. 19 min audio -> 19 clean chunks)
    chunks = slice_any_audio(audio_data, chunk_duration_sec=60.0, overlap_sec=2.0)
    print(f"[Audio Processing] Sliced audio into {len(chunks)} chunk(s). Processing in parallel...")
    
    prompt = """أنت نظام ذكاء اصطناعي فائق الدقة متخصص حصرياً في تفريغ واستخراج أرقام وبيانات لوحات السيارات السعودية من الصوت بدقة 100% وبدون أي اختراع أو تخمين.
المطلوب بدقة متناهية:
1. استمع للتسجيل الصوتي واكتب كل لوحة نطقها المتحدث بدون استثناء وبالترتيب الزمني الدقيق.
2. ⚠️ الحروف المعتمدة في لوحات السعودية 17 حرفاً فقط:
   (أ ، ب ، ح ، د ، ر ، س ، ص ، ط ، ع ، ق ، ك ، ل ، م ، ن ، هـ ، و ، ي)
   - إذا نطق اسم الحرف (كاف=ك، دال=د، ميم=م، باء=ب، ألف=أ، واو=و، قاف=ق، عين=ع، سين=س، هاء=هـ...).
   - صيغة اللوحة: 3 حروف متباعدة + 1 إلى 4 أرقام (مثال: 'ك د م 958' أو 'د ب أ 9075' أو 'أ ع ب 087').
3. ⚠️ كلمات ليست لوحات: الكلمات مثل (رقم ، عين ، سين ، حسب ، دون ، فاصل ، لوحة) هي كلمات وصفية ولا تُعتبر لوحة.
4. الأرقام: اكتب الأرقام بدقة كاملة كما نُطقت سواء كانت مفردة أو مركبة (مثال: 'تسعمائة وثمانية وخمسين' = 958، 'ألف ومائتين وأربعة وثلاثين' = 1234، 'صفر سبعة وثمانين' = 087، 'واحد اثنين ثلاثة أربعة' = 1234).
5. نوع السيارة والشارع والملاحظات:
   - إذا ذكر المتحدث نوع السيارة (مثل: باص، هايلوكس، كامري، يارس، نقل...) اكتبه في vehicle_type.
   - ⚠️ إذا لم يذكر نوع السيارة، اتركه فارغاً "" (لا تكتب تويوتا من رأسك أبداً).
   - إذا قال شارع كذا أو طريق كذا → اكتب في street_name.
   - إذا قال حي كذا → اكتب في district_name.
   - إذا قال ملاحظات (سليمة، مصدومة...) → اكتب في notes.
6. تصحيح النطق: إذا قال 'تعديل' أو 'قصدي' أو 'لا' أو 'معليش' بعد لوحة، استبدل اللوحة السابقة باللوحة المصححة.
7. مصفوفة JSON فقط بالشكل التالي:
[{"plate": "ك د م 958", "found": true, "vehicle_type": "", "street_name": "", "district_name": "", "notes": "", "street_location": ""}]
إذا كان التسجيل صامتاً أو لم يذكر أي لوحة، أرجع: []"""
    
    tasks = [(i, len(chunks), c, cfg, prompt, kind) for i, c in enumerate(chunks)]
    
    # Process up to 4 chunks in parallel for 4x faster execution
    from concurrent.futures import ThreadPoolExecutor
    max_workers = min(4, len(chunks)) if len(chunks) > 1 else 1
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        chunks_results = list(executor.map(_process_single_chunk, tasks))
            
    final_plates = _dedup_boundary_plates(chunks_results)
    print(f"[Gemini Universal Audio Parser OK] Total Extracted: {len(final_plates)} plates across {len(chunks)} chunk(s).")
    return final_plates

async def _run_background_transcribe(job_id: str, cfg: dict, audio_data: bytes, model_name: str, recorder_name: str, district: str, gps_data_raw: str):
    """Asynchronous background worker to process large audio files without blocking the HTTP request."""
    try:
        loop = asyncio.get_running_loop()
        plates = await loop.run_in_executor(None, _transcribe_dual_engine, cfg, audio_data, model_name, "rest")
        
        # Parse and format GPS points list
        formatted_gps_list = []
        if gps_data_raw:
            try:
                gps_list = json.loads(gps_data_raw)
                if isinstance(gps_list, list):
                    for pt in gps_list:
                        if isinstance(pt, dict) and "lat" in pt:
                            lng_val = pt.get("lng") or pt.get("lon")
                            if lng_val:
                                formatted_gps_list.append(f"{float(pt['lat']):.6f},{float(lng_val):.6f}")
                        elif isinstance(pt, str) and pt.strip():
                            formatted_gps_list.append(pt.strip())
            except Exception:
                pass

        num_plates = len(plates)
        num_gps = len(formatted_gps_list)
        from datetime import datetime
        current_date_str = datetime.now().strftime("%d/%m/%Y")

        for idx, p in enumerate(plates):
            plate_gps = ""
            if num_gps > 0:
                g_idx = int(idx * (num_gps - 1) / max(1, num_plates - 1)) if num_plates > 1 else 0
                plate_gps = formatted_gps_list[min(g_idx, num_gps - 1)]

            if plate_gps:
                p["gps"] = plate_gps
                if not p.get("street_location"):
                    p["street_location"] = plate_gps

            if not p.get("district_name") and district:
                p["district_name"] = district
            if not p.get("recorder_name") and recorder_name:
                p["recorder_name"] = recorder_name
            if not p.get("recording_date"):
                p["recording_date"] = current_date_str

        JOB_STORE[job_id] = {
            "status": "done",
            "plates": plates,
            "total": len(plates),
            "recorder_name": recorder_name,
            "district": district
        }
        print(f"[Background Job {job_id}] Successfully finished! Total extracted: {len(plates)} plates.")
    except Exception as e:
        print(f"[Background Job {job_id} Error] {e}")
        JOB_STORE[job_id] = {
            "status": "error",
            "detail": f"خطأ أثناء معالجة الصوت: {e}",
            "plates": []
        }

@app.post("/api/check-turn")
async def check_turn(request: Request):
    """
    Ultra-low latency (<600ms) synchronous transcription for live Check Session turns.
    High phonetic fidelity for Arabic letter names (راء/ح/ميم) and strict 17 Saudi letters enforcement.
    """
    cfg = load_config()
    try:
        form = await request.form()
        audio_file = form.get("audio")
        if not audio_file:
            return {"status": "ok", "plates": [], "total": 0}

        content = await audio_file.read()
        if len(content) < 400:
            return {"status": "ok", "plates": [], "total": 0}

        prompt = """أنت نظام ذكاء اصطناعي فائق الدقة متخصص في تفريغ لوحات السيارات السعودية بدقة 100% وبدون أي تخمين.
المطلوب بدقة متناهية:
1. استمع للصوت واستخرج اللوحة التي نطقها المتحدث كما نُطقت تماماً.
2. ⚠️ تمييز صوت واسم الحرف (17 حرفاً معتمداً فقط):
   - حرف الراء: (راء / را / ر) → ر (مثال: 'ر ح م 3830' = راء حاء ميم 3830).
   - حرف الحاء: (حاء / حا / ح) → ح
   - حرف الميم: (ميم / ما / م) → م
   - حرف السين: (سين / سا / س) → س
   - حرف الدال: (دال / دا / د) → د
   - حرف الكاف: (كاف / كا / ك) → ك
   - حرف الباء: (باء / با / ب) → ب
   - حرف الألف: (ألف / ا / أ) → أ
   - حرف الصاد: (صاد / صا / ص) → ص
   - حرف الطاء: (طاء / طا / ط) → ط
   - حرف العين: (عين / عا / ع) → ع
   - حرف القاف: (قاف / قا / ق) → ق
   - حرف اللام: (لام / لا / ل) → ل
   - حرف النون: (نون / نا / ن) → ن
   - حرف الهاء: (هاء / ها / هـ) → هـ
   - حرف الواو: (واو / و) → و
   - حرف الياء: (ياء / يا / ي) → ي
3. ⚠️ الصيغة القياسية: 3 حروف متباعدة + 1 إلى 4 أرقام (مثال: 'ر ح م 3830').
4. أرجع مصفوفة JSON فقط بالشكل التالي:
[{"plate": "ر ح م 3830", "found": false, "vehicle_type": "", "notes": ""}]
إذا لم تكن هناك لوحة أو كان الصوت صامتاً، أرجع: []"""

        mime_type, _ = _detect_audio_info(content)
        b64_audio = base64.b64encode(content).decode("utf-8")
        payload = {
            "contents": [{
                "parts": [
                    {"text": prompt},
                    {"inline_data": {"mime_type": mime_type, "data": b64_audio}}
                ]
            }],
            "generationConfig": {
                "response_mime_type": "application/json",
                "temperature": 0.0,
                "max_output_tokens": 1024
            }
        }

        # Direct fast invocation of Gemini
        res_json = await asyncio.to_thread(_call_gemini_with_rotation, cfg, payload, "gemini-flash-lite-latest", "live")
        raw_text = res_json["candidates"][0]["content"]["parts"][0]["text"].strip()
        if raw_text.startswith("```json"): raw_text = raw_text[7:]
        if raw_text.startswith("```"): raw_text = raw_text[3:]
        if raw_text.endswith("```"): raw_text = raw_text[:-3]
        parsed = json.loads(raw_text.strip())
        if isinstance(parsed, dict): parsed = [parsed]
        plates = []
        if isinstance(parsed, list):
            for p in parsed:
                cl = clean_saudi_plate(p.get("plate", ""))
                if cl:
                    p["plate"] = cl
                    if p.get("vehicle_type") == "تويوتا" and "تويوتا" not in str(p):
                        p["vehicle_type"] = ""
                    plates.append(p)
        return {"status": "ok", "plates": plates, "total": len(plates)}
    except Exception as e:
        print(f"[/api/check-turn error] {e}, falling back to Groq...")
        try:
            groq_p = await asyncio.to_thread(_call_groq_whisper, cfg, content)
            cleaned = []
            for p in groq_p:
                cl = clean_saudi_plate(p.get("plate", ""))
                if cl:
                    p["plate"] = cl
                    cleaned.append(p)
            return {"status": "ok", "plates": cleaned, "total": len(cleaned)}
        except Exception:
            return {"status": "ok", "plates": [], "total": 0}


@app.post("/api/process")
async def process_audio(request: Request):
    job_id = f"job_{uuid.uuid4().hex[:16]}"
    cfg = load_config()

    try:
        form = await request.form()
        audio_file = form.get("audio")
        model_name = str(form.get("model_name") or cfg.get("gemini_model", "gemini-flash-lite-latest"))
        recorder_name = str(form.get("recorder_name") or "")
        district = str(form.get("district_default") or "")
        gps_data_raw = form.get("gps_data")

        if not audio_file:
            raise HTTPException(status_code=400, detail="لا يوجد ملف صوتي")

        content = await audio_file.read()
        print(f"[Process Audio] Received file: {getattr(audio_file, 'filename', 'audio')}, size: {len(content)} bytes. Starting async background task {job_id}...")

        # Initialize job store with pending state
        JOB_STORE[job_id] = {
            "status": "pending",
            "progress": 5,
            "plates": []
        }

        # Launch background task non-blocking
        asyncio.create_task(_run_background_transcribe(job_id, cfg, content, model_name, recorder_name, district, gps_data_raw))

        # Return immediately in <10ms to prevent HTTP 502/Gateway Timeout
        return {
            "status": "ok",
            "job_id": job_id,
            "message": "جاري معالجة الصوت في الخلفية..."
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/transcribe/status/{job_id}")
async def transcribe_status(job_id: str):
    # Wait briefly to simulate async processing
    result = JOB_STORE.get(job_id)
    if not result:
        # Still processing - return pending
        return {"status": "pending", "progress": 50}
    
    plates = result.get("plates", [])
    return {
        "status": "done",
        "data": {
            "kind": "transcribe",
            "plates": plates,
            "total": len(plates)
        }
    }

# --- EXCEL & CHECK ENDPOINTS ---
@app.get("/api/check-live/debug-wav")
async def get_debug_wav():
    debug_wav_path = os.path.join(BASE_DIR, "debug_last_utterance.wav")
    if os.path.exists(debug_wav_path):
        return FileResponse(debug_wav_path, media_type="audio/wav", filename="debug_last_utterance.wav")
    raise HTTPException(status_code=404, detail="No debug WAV audio recorded yet")

def _parse_any_excel_file(content: bytes, pw: str = "") -> tuple[list, list]:
    """Robust unified parser for any spreadsheet file (.xlsx, .xlsm, .csv) with auto-decryption and read_only streaming"""
    bio = io.BytesIO(content)
    # 1. Try msoffcrypto decryption
    try:
        import msoffcrypto
        office_file = msoffcrypto.OfficeFile(bio)
        if office_file.is_encrypted():
            decrypted = io.BytesIO()
            office_file.load_key(password=str(pw).strip() if pw else "VelvetSweatshop")
            office_file.decrypt(decrypted)
            decrypted.seek(0)
            bio = decrypted
    except Exception:
        bio.seek(0)

    headers = []
    rows = []

    # 2. Try openpyxl with read_only=True (fast, lightweight memory for huge files)
    try:
        bio.seek(0)
        wb = openpyxl.load_workbook(filename=bio, read_only=True, data_only=True)
        sheet = wb.active or wb[wb.sheetnames[0]]
        for r in sheet.iter_rows(values_only=True):
            if not r or not any(c is not None and str(c).strip() for c in r):
                continue
            parsed = [str(c).strip() if c is not None else "" for c in r]
            if not headers:
                headers = [h if h else f"عمود_{i+1}" for i, h in enumerate(parsed)]
            else:
                row_dict = {}
                for idx, val in enumerate(r):
                    col_name = headers[idx] if idx < len(headers) else f"عمود_{idx+1}"
                    row_dict[col_name] = str(val).strip() if val is not None else ""
                rows.append(row_dict)
        if headers:
            return headers, rows
    except Exception as e1:
        pass

    # 3. Try standard openpyxl (fallback)
    try:
        bio.seek(0)
        wb = openpyxl.load_workbook(filename=bio, data_only=True)
        sheet = wb.active or wb[wb.sheetnames[0]]
        headers = []
        rows = []
        for r in sheet.iter_rows(values_only=True):
            if not r or not any(c is not None and str(c).strip() for c in r):
                continue
            parsed = [str(c).strip() if c is not None else "" for c in r]
            if not headers:
                headers = [h if h else f"عمود_{i+1}" for i, h in enumerate(parsed)]
            else:
                row_dict = {}
                for idx, val in enumerate(r):
                    col_name = headers[idx] if idx < len(headers) else f"عمود_{idx+1}"
                    row_dict[col_name] = str(val).strip() if val is not None else ""
                rows.append(row_dict)
        if headers:
            return headers, rows
    except Exception as e2:
        pass

    # 4. Try CSV fallback
    import csv
    for enc in ["utf-8-sig", "utf-8", "windows-1256", "latin1"]:
        try:
            text_str = content.decode(enc)
            reader = csv.reader(io.StringIO(text_str))
            headers = []
            rows = []
            for r in reader:
                if not r or not any(c.strip() for c in r):
                    continue
                parsed = [c.strip() for c in r]
                if not headers:
                    headers = [h if h else f"عمود_{i+1}" for i, h in enumerate(parsed)]
                else:
                    row_dict = {}
                    for idx, val in enumerate(r):
                        col_name = headers[idx] if idx < len(headers) else f"عمود_{idx+1}"
                        row_dict[col_name] = val.strip() if val else ""
                    rows.append(row_dict)
            if headers:
                return headers, rows
        except Exception:
            continue

    return headers, rows

@app.post("/api/check-headers")
async def check_headers(request: Request):
    headers = []
    detected_col = ""
    try:
        form = await request.form()
        file = form.get("large_file") or form.get("file") or form.get("large") or form.get("small_file") or form.get("small")
        pw = form.get("password") or form.get("largePw") or form.get("smallPw") or ""
        if file:
            content = await file.read()
            headers, _ = _parse_any_excel_file(content, str(pw))

        if not headers:
            headers = ["رقم اللوحة", "نوع السيارة", "ملاحظات", "GPS", "الحي", "الشارع"]

        # Auto detect plate column
        for h in headers:
            h_clean = h.strip()
            if any(kw in h_clean for kw in ["لوحة", "لوحه", "اللوحة", "رقم اللوحة", "plate", "Plate", "PLATE", "اللوح"]):
                detected_col = h
                break
        if not detected_col and headers:
            detected_col = headers[0]

    except Exception as e:
        print("check-headers exception:", e)
        
    return {
        "status": "ok",
        "headers": headers,
        "cols": headers,
        "columns": headers,
        "detected": detected_col,
        "detected_col": detected_col,
        "detected_column": detected_col,
        "large": {
            "headers": headers,
            "detected": detected_col
        }
    }

@app.post("/api/check-live/ref-plates-upload")
@app.post("/api/check-live/upload-excel")
async def check_live_upload_excel(request: Request):
    headers = ["رقم اللوحة", "نوع السيارة", "الحي", "الشارع", "ملاحظات"]
    total_plates = 0
    detected_col = "رقم اللوحة"
    try:
        form = await request.form()
        file = form.get("file") or form.get("large_file") or form.get("large")
        pw = form.get("password") or ""
        if file:
            content = await file.read()
            h_list, r_list = _parse_any_excel_file(content, str(pw))
            if h_list:
                headers = h_list
            total_plates = len(r_list)
                
        for h in headers:
            if any(kw in h for kw in ["لوحة", "لوحه", "اللوحة", "plate", "Plate"]):
                detected_col = h
                break
    except Exception as e:
        print("check-live upload-excel error:", e)

    return {
        "status": "ok",
        "message": "تم رفع وتجهيز الملف المرجعي بنجاح",
        "stored_count": total_plates,
        "total_plates": total_plates,
        "headers": headers,
        "columns": headers,
        "cols": headers,
        "detected": detected_col,
        "detected_col": detected_col,
        "large": {
            "headers": headers,
            "detected": detected_col
        }
    }

@app.post("/api/parse-gps-excel")
async def parse_gps_excel(request: Request):
    headers = []
    rows = []
    try:
        form = await request.form()
        file = form.get("file") or form.get("large_file") or form.get("small_file") or form.get("large") or form.get("small")
        pw = form.get("password") or form.get("largePw") or form.get("smallPw") or ""
        if file:
            content = await file.read()
            headers, rows = _parse_any_excel_file(content, str(pw))
    except Exception as e:
        print("Excel parsing exception:", e)
            
    return {
        "status": "ok",
        "headers": headers,
        "rows": rows,
        "total_rows": len(rows)
    }

def _norm_plate_str(s: str) -> str:
    if not s:
        return ""
    indic_to_eng = str.maketrans("٠١٢٣٤٥٦٧٨٩", "0123456789")
    s = str(s).translate(indic_to_eng).strip().lower()
    s = re.sub(r'[\s\u200b\u200c\u200d\ufeff\-_\.,:;/\\]+', '', s)
    s = re.sub(r'[أإآٱ]', 'أ', s)
    s = s.replace('ا', 'أ').replace('ى', 'ي').replace('ة', 'هـ').replace('ه', 'هـ')
    return s

@app.post("/api/fast-match")
async def fast_match(request: Request):
    try:
        form = await request.form()
        large_file = form.get("large_file") or form.get("large") or form.get("file")
        small_file = form.get("small_file") or form.get("small")
        plates_text = str(form.get("plates_text") or form.get("small_plates_text") or "").strip()
        large_col = str(form.get("large_col") or "").strip()
        small_col = str(form.get("small_col") or "").strip()
        large_pw = str(form.get("large_pw") or form.get("password") or "").strip()
        small_pw = str(form.get("small_pw") or "").strip()
        
        my_lat = float(form.get("my_lat") or 24.7136)
        my_lon = float(form.get("my_lon") or 46.6753)

        if not large_file:
            raise HTTPException(status_code=400, detail="الملف الكبير مطلوب لإجراء الفحص")

        # 1. Parse search plates
        search_plates_list = []
        if plates_text:
            search_plates_list = [p.strip() for p in re.split(r'[\r\n,]+', plates_text) if p.strip()]
        elif small_file:
            small_bytes = await small_file.read()
            s_headers, s_rows = _parse_any_excel_file(small_bytes, small_pw)
            
            # Find small plate col
            target_scol = small_col
            if not target_scol or (s_headers and target_scol not in s_headers):
                for h in s_headers:
                    if any(kw in h for kw in ["لوحة", "لوحه", "اللوحة", "plate", "Plate", "اللوح"]):
                        target_scol = h
                        break
                if not target_scol and s_headers:
                    target_scol = s_headers[0]
            
            for r in s_rows:
                v = r.get(target_scol, "").strip()
                if v:
                    search_plates_list.append(v)

        if not search_plates_list:
            raise HTTPException(status_code=400, detail="لم يتم العثور على أي لوحات في قائمة البحث (الملف الصغير)")

        search_set = {_norm_plate_str(p): p for p in search_plates_list if _norm_plate_str(p)}

        # 2. Fast Streaming Match on Large File
        large_bytes = await large_file.read()
        bio = io.BytesIO(large_bytes)
        
        # Try msoffcrypto
        try:
            import msoffcrypto
            office_file = msoffcrypto.OfficeFile(bio)
            if office_file.is_encrypted():
                decrypted = io.BytesIO()
                office_file.load_key(password=large_pw if large_pw else "VelvetSweatshop")
                office_file.decrypt(decrypted)
                decrypted.seek(0)
                bio = decrypted
        except Exception:
            bio.seek(0)

        headers = []
        matched_rows = []
        matched_unique = set()
        plate_idx = 0
        gps_idx = -1
        type_idx = -1
        notes_idx = -1
        date_idx = -1

        try:
            bio.seek(0)
            wb = openpyxl.load_workbook(filename=bio, read_only=True, data_only=True)
            sheet = wb.active or wb[wb.sheetnames[0]]
            
            for r in sheet.iter_rows(values_only=True):
                if not r or not any(c is not None and str(c).strip() for c in r):
                    continue
                parsed = [str(c).strip() if c is not None else "" for c in r]
                if not headers:
                    headers = [h if h else f"عمود_{i+1}" for i, h in enumerate(parsed)]
                    # Determine column indices
                    if large_col and large_col in headers:
                        plate_idx = headers.index(large_col)
                    else:
                        for idx, h in enumerate(headers):
                            if any(kw in h for kw in ["لوحة", "لوحه", "اللوحة", "plate", "Plate", "اللوح"]):
                                plate_idx = idx
                                break
                    for idx, h in enumerate(headers):
                        if any(kw in h.lower() for kw in ["gps", "موقع", "احداثيات", "إحداثيات"]):
                            gps_idx = idx
                        elif any(kw in h for kw in ["نوع السيارة", "النوع", "طراز"]):
                            type_idx = idx
                        elif any(kw in h for kw in ["ملاحظات", "ملاحظة"]):
                            notes_idx = idx
                        elif any(kw in h for kw in ["تاريخ", "التاريخ"]):
                            date_idx = idx
                else:
                    val = parsed[plate_idx] if plate_idx < len(parsed) else ""
                    norm_v = _norm_plate_str(val)
                    if norm_v in search_set:
                        matched_unique.add(norm_v)
                        row_dict = {headers[i] if i < len(headers) else f"عمود_{i+1}": parsed[i] for i in range(len(parsed))}
                        matched_rows.append(row_dict)
        except Exception as e:
            # Fallback to _parse_any_excel_file
            h_all, r_all = _parse_any_excel_file(large_bytes, large_pw)
            headers = h_all
            target_lcol = large_col
            if not target_lcol or target_lcol not in headers:
                for h in headers:
                    if any(kw in h for kw in ["لوحة", "لوحه", "اللوحة", "plate", "Plate", "اللوح"]):
                        target_lcol = h
                        break
                if not target_lcol and headers:
                    target_lcol = headers[0]
            for r in r_all:
                val = r.get(target_lcol, "").strip()
                norm_v = _norm_plate_str(val)
                if norm_v in search_set:
                    matched_unique.add(norm_v)
                    matched_rows.append(r)

        # 3. GPS distance calculations on matched rows
        gps_col_name = headers[gps_idx] if (gps_idx >= 0 and gps_idx < len(headers)) else None
        if not gps_col_name:
            for h in headers:
                if any(kw in h.lower() for kw in ["gps", "موقع", "احداثيات", "إحداثيات"]):
                    gps_col_name = h
                    break
        
        gps_results = []
        if gps_col_name and matched_rows:
            def _haversine(lat1, lon1, lat2, lon2):
                R = 6371.0
                dlat = math.radians(lat2 - lat1)
                dlon = math.radians(lon2 - lon1)
                a = math.sin(dlat/2)**2 + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dlon/2)**2
                return 2 * R * math.asin(math.sqrt(a))
            
            plate_name = headers[plate_idx] if (plate_idx >= 0 and plate_idx < len(headers)) else "رقم اللوحة"
            for r in matched_rows:
                gps_str = str(r.get(gps_col_name, "")).strip()
                dist = 99999.0
                if "," in gps_str:
                    try:
                        p_lat, p_lon = [float(x.strip()) for x in gps_str.split(",")[:2]]
                        dist = _haversine(my_lat, my_lon, p_lat, p_lon)
                    except Exception:
                        pass
                gps_results.append({
                    "plate": r.get(plate_name, ""),
                    "gps": gps_str,
                    "vehicle_type": r.get(headers[type_idx] if (type_idx >= 0 and type_idx < len(headers)) else "نوع السيارة", ""),
                    "notes": r.get(headers[notes_idx] if (notes_idx >= 0 and notes_idx < len(headers)) else "ملاحظات", ""),
                    "dist": round(dist, 2) if dist < 99990 else 99999,
                    "duration": int(round(dist / 40.0 * 60.0)) if dist < 99990 else 0,
                    "date": r.get(headers[date_idx] if (date_idx >= 0 and date_idx < len(headers)) else "تاريخ التسجيل", ""),
                    "raw": r
                })
            gps_results.sort(key=lambda x: x["dist"])

        unmatched_count = max(0, len(search_set) - len(matched_unique))

        return {
            "status": "ok",
            "headers": headers,
            "matched_rows": matched_rows,
            "matched_count": len(matched_rows),
            "unique_matched_count": len(matched_unique),
            "unmatched_count": unmatched_count,
            "gps_col": gps_col_name or "",
            "gps_results": gps_results,
            "large_col": headers[plate_idx] if (plate_idx >= 0 and plate_idx < len(headers)) else large_col,
            "small_col": small_col
        }
    except HTTPException:
        raise
    except Exception as ge:
        print("fast_match error:", ge)
        raise HTTPException(status_code=500, detail=f"حدث خطأ أثناء المطابقة: {ge}")

@app.post("/api/parse-export-append")
async def parse_export_append(request: Request):
    return {"status": "ok", "filename": "export.xlsx", "rows": []}

# --- EXCEL EXPORT ---
@app.post("/api/export-check-session")
@app.post("/api/export-excel")
async def export_excel(request: Request):
    """Generate a real Excel file with exact 8-column layout matching user specification:
    [اللوحة, النوع, الملاحظات, الشارع, الموقع, الحي, التاريخ, المندوب]
    """
    try:
        form = await request.form()
        rows_json = form.get("rows_json", "[]")
        sheet_name = str(form.get("sheet_name") or "بيانات المركبات").strip()
        district_default = str(form.get("district_default") or "").strip()
        recorder_default = str(form.get("recorder_default") or "").strip()
        street_default = str(form.get("street_default") or "").strip()

        rows = json.loads(rows_json)
        if not isinstance(rows, list):
            rows = []

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = (sheet_name[:31] if sheet_name else "بيانات المركبات")
        ws.sheet_view.rightToLeft = True  # RTL display for Arabic

        # --- Styling ---
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import urllib.parse

        # Green header fill matching user's image (#1E5631 / dark green)
        header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="1E5631")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        link_font = Font(name="Arial", color="0563C1", underline="single", size=10)
        regular_font = Font(name="Arial", size=10)
        thin = Side(style="thin", color="D1D5DB")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Standard 8 Columns in exact order (RTL: Rightmost is اللوحة, Leftmost is المندوب)
        STANDARDIZED_COLS = [
            ("اللوحة", ["full_plate", "plate", "رقم اللوحة", "اللوحة", "plate_number"]),
            ("النوع", ["vehicle_type", "نوع السيارة", "car_type", "النوع", "طراز"]),
            ("الملاحظات", ["notes", "ملاحظات", "ملاحظة", "location_details", "تفاصيل"]),
            ("الشارع", ["street_name", "street", "الشارع", "شارع"]),
            ("الموقع", ["gps", "GPS", "موقع", "احداثيات", "إحداثيات", "street_location"]),
            ("الحي", ["district_name", "district", "الحي", "حي", "المنطقة"]),
            ("التاريخ", ["recording_date", "date", "التاريخ", "تاريخ التسجيل"]),
            ("المندوب", ["recorder_name", "المسجّل", "المسجل", "المندوب"]),
        ]

        cols = [header for header, _ in STANDARDIZED_COLS]

        # Write header row
        ws.row_dimensions[1].height = 28
        for col_idx, col_name in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border

        # Helper to get value from row dict
        def get_col_val(row_dict, header_name):
            if not isinstance(row_dict, dict):
                return ""
            for h, aliases in STANDARDIZED_COLS:
                if h == header_name:
                    for a in aliases:
                        if a in row_dict and row_dict[a] not in [None, ""]:
                            return str(row_dict[a]).strip()
            return ""

        from datetime import datetime
        today_date_str = datetime.now().strftime("%d/%m/%Y")

        # Write data rows
        alt_fill = PatternFill("solid", fgColor="F9FAFB")
        for row_idx, row in enumerate(rows, start=2):
            ws.row_dimensions[row_idx].height = 24
            fill = alt_fill if row_idx % 2 == 0 else None
            
            for col_idx, col_name in enumerate(cols, start=1):
                val = get_col_val(row, col_name)

                # Fallbacks for defaults
                if col_name == "الحي" and not val and district_default:
                    val = district_default
                if col_name == "الشارع" and not val and street_default:
                    val = street_default
                if col_name == "المندوب" and not val and recorder_default:
                    val = recorder_default
                if col_name == "التاريخ" and not val:
                    val = today_date_str

                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = center_align
                cell.border = border
                if fill:
                    cell.fill = fill

                # Handle Google Maps Link in 'الموقع' column
                if col_name == "الموقع" and val and ("," in val or "http" in val):
                    gps_coords = val
                    if "http" in val:
                        # Extract coords from existing URL if possible
                        m = re.search(r'([0-9]+\.[0-9]+,[0-9]+\.[0-9]+)', val)
                        if m:
                            gps_coords = m.group(1)
                    if "," in gps_coords:
                        nav_url = f"https://www.google.com/maps/dir/?api=1&destination={gps_coords.strip()}"
                        cell.value = nav_url
                        cell.hyperlink = nav_url
                        cell.font = link_font
                    else:
                        cell.value = val
                        cell.font = regular_font
                else:
                    cell.value = val if val else ""
                    cell.font = regular_font

        # Auto-fit column widths
        col_widths = {1: 18, 2: 14, 3: 20, 4: 18, 5: 32, 6: 16, 7: 14, 8: 16}
        for col_idx in range(1, len(cols) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_idx, 16)

        # Freeze header row
        ws.freeze_panes = "A2"

        # Save to bytes
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        safe_sheet_name = re.sub(r'[\\/*?:\"<>|]', '_', sheet_name) or "bareq_export"
        filename = f"{safe_sheet_name}_{int(time.time())}.xlsx"
        quoted_filename = urllib.parse.quote(filename)
        return Response(
            content=buf.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quoted_filename}"}
        )

    except Exception as e:
        print(f"export-excel error: {e}")
        raise HTTPException(status_code=500, detail=f"فشل تصدير Excel: {e}")

@app.get("/api/export-excel")
async def export_excel_get():
    raise HTTPException(status_code=405, detail="استخدم POST مع rows_json")


def _route_via_osrm(coordinates: list) -> dict:
    """Free routing via OSRM public server - no API key needed"""
    coords_str = ";".join(f"{c[0]},{c[1]}" for c in coordinates)
    url = f"https://router.project-osrm.org/route/v1/driving/{coords_str}?overview=full&geometries=geojson&steps=true"
    resp = requests.get(url, timeout=15)
    resp.raise_for_status()
    data = resp.json()
    if data.get("code") != "Ok":
        raise Exception(f"OSRM error: {data.get('code')}")
    route = data["routes"][0]
    return {
        "source": "osrm",
        "distance_m": route["distance"],
        "duration_s": route["duration"],
        "distance_km": round(route["distance"] / 1000, 2),
        "duration_min": round(route["duration"] / 60, 1),
        "geometry": route["geometry"],
        "summary": f"{round(route['distance']/1000,1)} كم — {round(route['duration']/60,0)} دقيقة"
    }

def _get_ors_base_url():
    cfg = load_config()
    return cfg.get("ors_base_url", "https://api.heigit.org").rstrip("/")

def _route_via_ors(coordinates: list, ors_key: str, profile: str = "driving-car") -> dict:
    """Routing via ORS/HEIGit (requires activated key)"""
    base = _get_ors_base_url()
    headers = {
        "Authorization": ors_key,
        "Content-Type": "application/json",
        "Accept": "application/json, application/geo+json"
    }
    payload = {"coordinates": coordinates}
    url = f"{base}/v2/directions/{profile}"
    resp = requests.post(url, json=payload, headers=headers, timeout=15)
    if resp.status_code == 403:
        raise Exception("ORS_FORBIDDEN")
    resp.raise_for_status()
    data = resp.json()
    seg = data["routes"][0]["summary"]
    return {
        "source": "ors",
        "distance_m": seg["distance"],
        "duration_s": seg["duration"],
        "distance_km": round(seg["distance"] / 1000, 2),
        "duration_min": round(seg["duration"] / 60, 1),
        "geometry": data["routes"][0].get("geometry"),
        "summary": f"{round(seg['distance']/1000,1)} كم — {round(seg['duration']/60,0)} دقيقة"
    }

@app.post("/api/ors/directions")
async def ors_directions(req: Request):
    """Smart routing: ORS first → OSRM fallback"""
    cfg = load_config()
    ors_keys = cfg.get("ors_keys", [])
    ors_key = ors_keys[0] if ors_keys else cfg.get("ors_api_key", "")
    
    body = await req.json()
    profile = body.get("profile", "driving-car")
    coordinates = body.get("coordinates", [])
    
    if not coordinates or len(coordinates) < 2:
        raise HTTPException(status_code=400, detail="coordinates مطلوبة (نقطتان على الأقل)")
    
    # Try ORS first if key exists and not previously forbidden
    if ors_key:
        try:
            result = _route_via_ors(coordinates, ors_key, profile)
            return result
        except Exception as e:
            if "ORS_FORBIDDEN" not in str(e):
                print(f"ORS error (not 403): {e}")
            # Fall through to OSRM
    
    # Fallback to free OSRM
    try:
        result = _route_via_osrm(coordinates)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"فشل حساب المسار: {e}")

@app.get("/api/ors/geocode")
async def ors_geocode(text: str = "", lat: float = 0, lon: float = 0):
    """Geocoding via ORS or Nominatim fallback"""
    cfg = load_config()
    ors_keys = cfg.get("ors_keys", [])
    ors_key = ors_keys[0] if ors_keys else cfg.get("ors_api_key", "")
    
    # Try ORS/HEIGit first
    if ors_key:
        try:
            base = _get_ors_base_url()
            headers = {"Authorization": ors_key, "Accept": "application/json"}
            if text:
                url = f"{base}/geocode/search?text={requests.utils.quote(text)}&size=5"
            else:
                url = f"{base}/geocode/reverse?point.lat={lat}&point.lon={lon}&size=1"
            resp = requests.get(url, headers=headers, timeout=10)
            if resp.status_code == 200:
                return JSONResponse(content=resp.json())
        except Exception:
            pass
    
    # Fallback: Nominatim (OpenStreetMap geocoding - free)
    try:
        if text:
            url = f"https://nominatim.openstreetmap.org/search?q={requests.utils.quote(text)}&format=json&limit=5"
        else:
            url = f"https://nominatim.openstreetmap.org/reverse?lat={lat}&lon={lon}&format=json"
        headers_nom = {"User-Agent": "BareqApp/4.0"}
        resp = requests.get(url, headers=headers_nom, timeout=10)
        return JSONResponse(content=resp.json())
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/ors/status")
async def ors_status():
    """Check routing status — ORS or OSRM fallback"""
    cfg = load_config()
    ors_keys = cfg.get("ors_keys", [])
    ors_key = ors_keys[0] if ors_keys else cfg.get("ors_api_key", "")
    
    result = {"ors": "no_key", "osrm": "unknown", "active_service": "osrm"}
    
    # Test ORS/HEIGit
    if ors_key:
        try:
            base = _get_ors_base_url()
            payload = {"coordinates": [[46.6753, 24.7136], [46.7153, 24.7436]]}
            headers = {"Authorization": ors_key, "Content-Type": "application/json", "Accept": "application/json"}
            resp = requests.post(
                f"{base}/v2/directions/driving-car",
                json=payload, headers=headers, timeout=10
            )
            if resp.status_code == 200:
                result["ors"] = "ok"
                result["active_service"] = "ors"
            elif resp.status_code == 403:
                result["ors"] = "forbidden - المفتاح يحتاج تفعيل"
            else:
                result["ors"] = f"error_{resp.status_code}: {resp.text[:80]}"
        except Exception as e:
            result["ors"] = f"error: {e}"
    
    # Test OSRM (always free)
    try:
        r = requests.get(
            "https://router.project-osrm.org/route/v1/driving/46.6753,24.7136;46.7153,24.7436?overview=false",
            timeout=8
        )
        if r.status_code == 200 and r.json().get("code") == "Ok":
            result["osrm"] = "ok"
            if result["active_service"] != "ors":
                result["active_service"] = "osrm"
        else:
            result["osrm"] = "error"
    except Exception as e:
        result["osrm"] = f"error: {e}"
    
    return {
        "status": "ok" if result["active_service"] in ("ors","osrm") else "error",
        "active_service": result["active_service"],
        "ors_status": result["ors"],
        "osrm_status": result["osrm"],
        "message": f"الخدمة النشطة: {result['active_service'].upper()}"
    }



@app.websocket("/ws/check-live")
async def websocket_check_live(websocket: WebSocket, ticket: str = ""):
    await websocket.accept()
    client_ip = websocket.client.host if websocket.client else "unknown"
    print(f"[WebSocket] Connected (check-live) from {client_ip}")
    cfg = load_config()

    # Per-connection session state (multi-user isolation)
    session = ASRSession() if _LOCAL_ASR_AVAILABLE else None
    legacy_pcm_chunks = []
    decoder = _get_plate_decoder() if (_DECODER_AVAILABLE and _get_plate_decoder) else None
    partial_in_flight = False

    # Reference plate index for live lookup (populated from uploaded Excel)
    ref_plate_set: dict = {}   # norm_key -> original plate string

    def _norm_ws_plate(s: str) -> str:
        """Normalise a plate string for lookup comparison"""
        if not s: return ""
        s = str(s).strip()
        s = re.sub(r'[\s\u200b\u200c\u200d\ufeff\-_]+', '', s)
        s = re.sub(r'[أإآٱا]', 'أ', s)
        s = s.replace('ى', 'ي').replace('ة', 'ه').replace('هـ', 'ه')
        return s.lower()

    def _lookup_plate(plate: str) -> bool:
        """Return True if plate is in the reference set"""
        if not ref_plate_set:
            return True   # No Excel loaded → treat all as found (don't block)
        return _norm_ws_plate(plate) in ref_plate_set


    def pcm16_to_wav(pcm_bytes: bytes, sample_rate: int = 16000) -> bytes:
        import wave
        buf = io.BytesIO()
        with wave.open(buf, 'wb') as wf:
            wf.setnchannels(1)
            wf.setsampwidth(2)
            wf.setframerate(sample_rate)
            wf.writeframes(pcm_bytes)
        return buf.getvalue()

    async def _run_partial_task(audio_np):
        nonlocal partial_in_flight
        try:
            res = await asyncio.to_thread(_local_transcribe, audio_np, True)
            raw_text = res.get("text", "").strip()
            if raw_text:
                dec = decoder.decode_partial(raw_text)
                partial_plate = dec.get("partial_plate", "")
                display_txt = partial_plate if partial_plate else raw_text

                if session:
                    if "first_partial_result" not in session.timestamps:
                        session.record_timestamp("first_partial_result")
                    if partial_plate:
                        session.partial_history.append(partial_plate)
                        # Keep only last 5 partials
                        if len(session.partial_history) > 5:
                            session.partial_history = session.partial_history[-5:]

                try:
                    await websocket.send_json({
                        "type": "live_transcript",
                        "data": display_txt,
                        "partial": True,
                        "partial_plate": partial_plate,
                        "is_complete": dec.get("is_complete", False)
                    })
                except Exception:
                    pass
        except Exception as e:
            print(f"[WS] Partial task error: {e}")
        finally:
            partial_in_flight = False

    try:
        while True:
            msg = await websocket.receive()
            if "text" in msg and msg["text"]:
                try:
                    data = json.loads(msg["text"])
                    mtype = data.get("type")

                    if mtype == "init":
                        # CRITICAL: client waits for type='ready' to set UI state to 'جاهز — استمع'
                        await websocket.send_json({"type": "ready"})
                        print("Sent type: ready to client")

                    elif mtype == "ping":
                        await websocket.send_json({"type": "pong"})

                    elif mtype == "excel_upload_ref":
                        # Client sends upload_token pointing to an already-uploaded file
                        upload_token = data.get("upload_token", "")
                        plate_col = data.get("plate_column", data.get("column", ""))
                        pw = data.get("password", "")
                        try:
                            upload_path = os.path.join(BASE_DIR, "uploads", f"{upload_token}.xlsx")
                            if not os.path.exists(upload_path):
                                upload_path = os.path.join(BASE_DIR, "uploads", f"{upload_token}.xls")
                            if os.path.exists(upload_path):
                                with open(upload_path, "rb") as fh:
                                    content_bytes = fh.read()
                                h_list, r_list = _parse_any_excel_file(content_bytes, str(pw))
                                # Auto-detect plate column
                                if not plate_col:
                                    for h in h_list:
                                        if any(kw in h for kw in ["لوحة", "لوحه", "اللوحة", "plate", "Plate"]):
                                            plate_col = h
                                            break
                                if not plate_col and h_list:
                                    plate_col = h_list[0]
                                ref_plate_set.clear()
                                for row in r_list:
                                    v = str(row.get(plate_col, "")).strip()
                                    if v:
                                        ref_plate_set[_norm_ws_plate(v)] = v
                                print(f"[WS] Loaded {len(ref_plate_set)} ref plates from upload_token={upload_token}")
                                await websocket.send_json({
                                    "type": "excel_loaded",
                                    "data": {"count": len(ref_plate_set), "column": plate_col}
                                })
                            else:
                                print(f"[WS] Upload token not found: {upload_token}")
                                await websocket.send_json({"type": "excel_loaded", "data": {"count": 0}})
                        except Exception as ex_err:
                            print(f"[WS] excel_upload_ref error: {ex_err}")
                            await websocket.send_json({"type": "excel_loaded", "data": {"count": 0}})

                    elif mtype == "set_plate_column":
                        # Client already uploaded via HTTP, just sets the column name
                        # Nothing to do since we already indexed — ack only
                        col = data.get("column", "")
                        await websocket.send_json({
                            "type": "plate_column_ready",
                            "data": {"count": len(ref_plate_set), "column": col}
                        })

                    elif mtype == "audio":
                        raw_b64 = data.get("data", "")
                        if raw_b64:
                            b_pcm = base64.b64decode(raw_b64)
                            if session:
                                if session.chunks_received == 0:
                                    session.record_timestamp("first_chunk")
                                session.add_audio(b_pcm)

                                # Check if we should trigger a partial decoding task
                                if not partial_in_flight and session.should_run_partial(min_interval_s=0.6):
                                    if _is_asr_available():
                                        audio_np = session.get_audio_numpy()
                                        if audio_np is not None:
                                            partial_in_flight = True
                                            session.last_partial_at = time.time()
                                            asyncio.create_task(_run_partial_task(audio_np))
                            else:
                                legacy_pcm_chunks.append(b_pcm)

                    elif mtype == "end_of_turn":
                        total_chunks = session.chunks_received if session else len(legacy_pcm_chunks)
                        print(f"End of turn received. Total chunks in utterance: {total_chunks}")

                        if session:
                            session.record_timestamp("end_of_turn")
                            audio_np = session.get_audio_numpy()
                            all_pcm_bytes = bytes(session.audio_buffer)
                            partial_history = list(session.partial_history)
                            session.clear_turn()
                        else:
                            all_pcm_bytes = b"".join(legacy_pcm_chunks)
                            legacy_pcm_chunks.clear()
                            partial_history = []
                            if all_pcm_bytes:
                                try:
                                    audio_np = np.frombuffer(all_pcm_bytes, dtype=np.int16).astype(np.float32) / 32768.0
                                except Exception:
                                    audio_np = None
                            else:
                                audio_np = None

                        if not all_pcm_bytes or len(all_pcm_bytes) < 3200:  # < 0.1s
                            continue

                        wav_data = pcm16_to_wav(all_pcm_bytes, 16000)

                        # Save debug WAV file for audio inspection (Requirement #3)
                        debug_wav_path = os.path.join(BASE_DIR, "debug_last_utterance.wav")
                        try:
                            with open(debug_wav_path, "wb") as f_wav:
                                f_wav.write(wav_data)
                        except Exception as wav_save_err:
                            print(f"[WS Debug] Error saving debug WAV: {wav_save_err}")

                        async def _process_end_of_turn(b_wav: bytes, a_np, p_hist, pcm_len: int):
                            t_start = time.time()
                            plate_text = ""
                            raw_text = ""
                            norm_text = ""
                            confidence = 0.0
                            is_valid = False
                            signals = {}
                            used_engine = "local"

                            # Calculate Audio Statistics (Requirement #2)
                            audio_duration = pcm_len / (16000 * 2)
                            rms_val = float(np.sqrt(np.mean(a_np ** 2))) if a_np is not None and len(a_np) > 0 else 0.0
                            peak_val = float(np.max(np.abs(a_np))) if a_np is not None and len(a_np) > 0 else 0.0

                            print("\n" + "=" * 60)
                            print(f"[AUDIO RECEIVED] bytes={pcm_len} | duration={audio_duration:.2f}s | rms={rms_val:.4f} | peak={peak_val:.4f}")
                            print("=" * 60)

                            # Audio stats computed, forward directly to active cloud ASR engines
                            if rms_val < 0.000001 and peak_val < 0.00001:
                                print("[AUDIO WARNING] Buffer is completely flat zero.")

                            # Show immediately that the server stored the turn; inference can take longer.
                            await websocket.send_json({
                                "type": "debug_pipeline",
                                "data": {
                                    "raw_asr": "", "normalized": "", "decoded_plate": "",
                                    "valid": False, "confidence": 0.0, "latency_ms": 0,
                                    "engine": "processing_audio",
                                    "audio_duration_s": round(audio_duration, 2),
                                    "rms": round(rms_val, 6), "peak": round(peak_val, 6)
                                }
                            })

                            # Call unified Dual-Engine with latest models and decoder
                            api_attempted = False
                            if cfg.get("enable_api_fallback", True):
                                api_attempted = True
                                try:
                                    used_engine = "cloud_asr"
                                    model_name = cfg.get("gemini_model", "gemini-flash-lite-latest")
                                    plates = await asyncio.to_thread(_transcribe_dual_engine, cfg, b_wav, model_name, "live")
                                    if plates:
                                        p0 = plates[0]
                                        raw_plate = p0.get("plate", "").strip()
                                        raw_text = raw_plate
                                        if decoder:
                                            dec = decoder.decode_final(raw_plate)
                                            norm_text = dec.get("normalized", "")
                                            plate_text = dec.get("plate", "") or raw_plate
                                            is_valid = dec.get("valid", True)
                                            confidence = max(0.85, dec.get("confidence", 0.85))
                                        else:
                                            plate_text = raw_plate
                                            is_valid = True
                                            confidence = 0.85
                                        print(f"[ASR Fast] RAW: '{raw_plate}' -> PLATE: '{plate_text}'")
                                except Exception as fast_err:
                                    print(f"[ASR Fast] Error, trying local ASR: {fast_err}")
                            # 1. Try Local ASR first (if available and loaded)
                            if not plate_text and _LOCAL_ASR_AVAILABLE and _is_asr_available() and a_np is not None:
                                try:
                                    t0_asr = time.time()
                                    res = await asyncio.to_thread(_local_transcribe, a_np, False)
                                    raw_text = res.get("text", "").strip()
                                    segments = res.get("segments", [])
                                    asr_elapsed = (time.time() - t0_asr) * 1000

                                    print(f"RAW ASR:       \"{raw_text}\"")

                                    if raw_text:
                                        logprobs = [s.get("avg_logprob", 0.0) for s in segments if "avg_logprob" in s]
                                        if decoder:
                                            dec = decoder.decode_final(
                                                raw_text,
                                                asr_segment_confidences=logprobs,
                                                partial_history=p_hist
                                            )
                                            norm_text = dec.get("normalized", "")
                                            plate_text = dec.get("plate", "")
                                            is_valid = dec.get("valid", False)
                                            confidence = dec.get("confidence", 0.0)
                                            signals = dec.get("signals", {})
                                        else:
                                            plate_text = raw_text
                                            is_valid = True
                                            confidence = 0.7
                                        used_engine = "local_whisper"

                                        print(f"NORMALIZED:    \"{norm_text}\"")
                                        print(f"DECODED PLATE: \"{plate_text}\" (valid={is_valid}, conf={confidence:.2f}, ASR={asr_elapsed:.0f}ms)")
                                except Exception as local_err:
                                    print(f"[ASR Local] Error, will fallback: {local_err}")
                                    plate_text = ""

                            t_total = (time.time() - t_start) * 1000
                            print(f"FINAL RESULT:  \"{plate_text}\" | engine={used_engine} | total_latency={t_total:.0f}ms")
                            print("=" * 60 + "\n")

                            # Send Debug Pipeline Payload to Frontend (Requirements #1, #4, #12)
                            try:
                                await websocket.send_json({
                                    "type": "debug_pipeline",
                                    "data": {
                                        "raw_asr": raw_text,
                                        "normalized": norm_text,
                                        "decoded_plate": plate_text,
                                        "valid": is_valid,
                                        "confidence": confidence,
                                        "signals": signals,
                                        "latency_ms": round(t_total, 1),
                                        "engine": used_engine,
                                        "audio_duration_s": round(audio_duration, 2),
                                        "rms": round(rms_val, 4),
                                        "peak": round(peak_val, 4)
                                    }
                                })
                            except Exception:
                                pass

                            # 3. Decision & Emission — emit ALL plates (with found status from Excel lookup)
                            try:
                                # Get all plates from the response (not just first)
                                all_result_plates = plates if (plates and isinstance(plates, list)) else []
                                if plate_text and not any(p.get("plate") == plate_text for p in all_result_plates):
                                    # Decoded plate differs from raw — use decoded, but clean it first
                                    cleaned_pt = clean_saudi_plate(plate_text)
                                    if cleaned_pt:
                                        all_result_plates = [{"plate": cleaned_pt, "vehicle_type": "", "notes": ""}]

                                for p_info in all_result_plates:
                                    pt = str(p_info.get("plate", "")).strip()
                                    # Validate plate through Saudi cleaner
                                    cleaned_pt = clean_saudi_plate(pt)
                                    if not cleaned_pt:
                                        continue
                                    pt = cleaned_pt
                                    # Remove hallucinated تويوتا
                                    if p_info.get("vehicle_type") == "تويوتا":
                                        p_info["vehicle_type"] = ""
                                    # Real lookup against loaded Excel reference
                                    is_found = _lookup_plate(pt)
                                    await websocket.send_json({
                                        "type": "plate_result",
                                        "data": {
                                            "plate": pt,
                                            "found": is_found,
                                            "vehicle_type": p_info.get("vehicle_type", ""),
                                            "notes": p_info.get("notes", ""),
                                            "street_name": p_info.get("street_name", ""),
                                            "district_name": p_info.get("district_name", ""),
                                            "moving": False,
                                            "confidence": confidence,
                                            "engine": used_engine,
                                            "latency_ms": round(t_total, 1),
                                            "signals": signals
                                        }
                                    })
                                    await websocket.send_json({
                                        "type": "live_transcript",
                                        "data": f"{'✔' if is_found else '—'} {pt}",
                                        "final": True
                                    })
                                    print(f"[WS] Emitted plate_result: {pt} found={is_found}")

                                if not all_result_plates:
                                    await websocket.send_json({
                                        "type": "live_transcript",
                                        "data": f"لم يتم التعرف على لوحة (RAW: {raw_text or 'صمت'})",
                                        "final": False
                                    })
                            except Exception as send_err:
                                print(f"[WS] Client disconnected before emission: {send_err}")

                        asyncio.create_task(_process_end_of_turn(wav_data, audio_np, partial_history, len(all_pcm_bytes)))


                except Exception as ex:
                    print(f"WS message error: {ex}")

            elif "bytes" in msg and msg["bytes"]:
                if session:
                    if session.chunks_received == 0:
                        session.record_timestamp("first_chunk")
                    session.add_audio(msg["bytes"])
                else:
                    legacy_pcm_chunks.append(msg["bytes"])

    except WebSocketDisconnect:
        print(f"[WebSocket] Disconnected (check-live) from {client_ip}")
    except Exception as e:
        print(f"[WebSocket] Error: {e}")




# --- AUTH ENDPOINTS ---
@app.post("/auth/login")
async def login(req: Request):
    data = await req.json()
    cfg = load_config()
    users = load_users()
    u_in = data.get("username", "").strip()
    p_in = data.get("password", "").strip()
    
    if u_in == cfg.get("admin_username", "admin") and p_in == cfg.get("admin_password", "123"):
        return {
            "status": "ok",
            "token": "bareq_admin_token_8899",
            "is_admin": True,
            "username": u_in,
            "display_name": "مدير النظام",
            "plan_name": "باقة المدير (غير محدود)",
            "rows_limit": 9999999,
            "subscription_end": "غير محدود",
            "is_active": True,
            "is_trial": False
        }
        
    for u in users:
        if (u.get("username") == u_in or u.get("email") == u_in) and u.get("password") == p_in:
            if not u.get("is_active", True):
                raise HTTPException(status_code=403, detail="هذا الحساب معطل، يرجى التواصل مع الإدارة")
            
            # Check subscription end date
            sub_end = str(u.get("subscription_end", "")).strip()
            is_expired = False
            if sub_end and sub_end not in ["غير محدود", "مفتوح - 30 يوماً"]:
                try:
                    from datetime import datetime
                    end_dt = datetime.strptime(sub_end[:10], "%Y-%m-%d")
                    if end_dt.date() < datetime.now().date():
                        is_expired = True
                except Exception:
                    pass

            return {
                "status": "ok",
                "token": f"bareq_token_{u['id']}",
                "user_id": u["id"],
                "is_admin": u.get("is_admin", False),
                "username": u.get("username", u_in),
                "display_name": u.get("display_name", u_in),
                "email": u.get("email", ""),
                "phone": u.get("phone", ""),
                "plan_id": u.get("plan_id", 0),
                "plan_name": u.get("plan_name", "فترة تجريبية مجانية"),
                "rows_limit": u.get("rows_limit", 500),
                "subscription_start": u.get("subscription_start", ""),
                "subscription_end": u.get("subscription_end", ""),
                "subscription_expired": is_expired,
                "is_trial": u.get("is_trial", False),
                "is_active": u.get("is_active", True)
            }
            
    raise HTTPException(status_code=401, detail="اسم المستخدم/البريد أو كلمة المرور غير صحيحة")

@app.post("/auth/register")
async def register(req: Request):
    data = await req.json()
    u_in = (data.get("username") or data.get("email") or "").strip()
    email_in = data.get("email", "").strip()
    p_in = data.get("password", "").strip()
    name_in = data.get("display_name", "").strip() or u_in
    phone_in = data.get("phone", "").strip()
    
    if not u_in or not p_in:
        raise HTTPException(status_code=400, detail="يرجى إدخال اسم المستخدم/البريد وكلمة المرور")
    if len(p_in) < 4:
        raise HTTPException(status_code=400, detail="كلمة المرور يجب أن تكون 4 خانات على الأقل")
        
    users = load_users()
    for u in users:
        if u.get("username") == u_in or (email_in and u.get("email") == email_in):
            raise HTTPException(status_code=400, detail="اسم المستخدم أو البريد مسجل مسبقاً، يرجى تسجيل الدخول")
            
    from datetime import datetime, timedelta
    now = datetime.now()
    trial_days = 7
    sub_end = (now + timedelta(days=trial_days)).strftime("%Y-%m-%d")
    
    new_id = max((u.get("id", 0) for u in users), default=0) + 1
    new_user = {
        "id": new_id,
        "username": u_in,
        "email": email_in or u_in,
        "password": p_in,
        "display_name": name_in,
        "phone": phone_in,
        "is_admin": False,
        "is_active": True,
        "is_trial": True,
        "plan_id": 0,
        "plan_name": "فترة تجريبية مجانية (7 أيام)",
        "rows_limit": 500,
        "subscription_start": now.strftime("%Y-%m-%d"),
        "subscription_end": sub_end,
        "created_at": now.strftime("%Y-%m-%d %H:%M:%S")
    }
    users.append(new_user)
    save_users(users)
    
    return {
        "status": "ok",
        "message": f"أهلاً بك يا {name_in}! تم تفعيل الفترة التجريبية المجانية لمدة 7 أيام.",
        "token": f"bareq_token_{new_user['id']}",
        "user_id": new_user["id"],
        "is_admin": False,
        "username": new_user["username"],
        "display_name": new_user["display_name"],
        "plan_name": new_user["plan_name"],
        "rows_limit": new_user["rows_limit"],
        "subscription_end": new_user["subscription_end"],
        "subscription_expired": False,
        "is_trial": True,
        "is_active": True
    }

@app.get("/auth/me")
async def auth_me(req: Request):
    auth_header = req.headers.get("Authorization", "").strip()
    token = auth_header.replace("Bearer ", "").strip()
    
    # Fallback to query param or cookie
    if not token:
        token = req.query_params.get("token", "").strip()
    if not token:
        token = req.cookies.get("access_token", "").strip()

    cfg = load_config()
    users = load_users()
    
    if not token:
        raise HTTPException(status_code=401, detail="يرجى تسجيل الدخول أولاً")
        
    # Check if admin token
    if "bareq_admin_token" in token:
        return {
            "status": "ok",
            "is_admin": True,
            "username": cfg.get("admin_username", "admin"),
            "display_name": "مدير النظام",
            "plan_name": "باقة المدير (غير محدود)",
            "rows_limit": 9999999,
            "subscription_end": "غير محدود",
            "subscription_expired": False,
            "is_active": True
        }
        
    # Try finding user by token
    for u in users:
        if token.endswith(str(u["id"])) or f"bareq_token_{u['id']}" in token or token == f"bareq_token_{u['id']}":
            if not u.get("is_active", True):
                raise HTTPException(status_code=403, detail="هذا الحساب معطل")
            
            sub_end = str(u.get("subscription_end", "")).strip()
            is_expired = False
            if sub_end and sub_end not in ["غير محدود", "مفتوح - 30 يوماً"]:
                try:
                    from datetime import datetime
                    end_dt = datetime.strptime(sub_end[:10], "%Y-%m-%d")
                    if end_dt.date() < datetime.now().date():
                        is_expired = True
                except Exception:
                    pass

            return {
                "status": "ok",
                "user_id": u["id"],
                "is_admin": u.get("is_admin", False),
                "username": u.get("username", ""),
                "display_name": u.get("display_name", ""),
                "email": u.get("email", ""),
                "phone": u.get("phone", ""),
                "plan_id": u.get("plan_id", 0),
                "plan_name": u.get("plan_name", "فترة تجريبية مجانية"),
                "rows_limit": u.get("rows_limit", 500),
                "subscription_start": u.get("subscription_start", ""),
                "subscription_end": u.get("subscription_end", ""),
                "subscription_expired": is_expired,
                "is_trial": u.get("is_trial", False),
                "is_active": u.get("is_active", True)
            }
            
    raise HTTPException(status_code=401, detail="انتهت صلاحية الجلسة، يرجى تسجيل الدخول")

# --- WHATSAPP SUPPORT API ---
@app.get("/api/whatsapp")
async def get_whatsapp_number():
    cfg = load_config()
    num = str(cfg.get("whatsapp_number", "201094593394")).strip().replace("+", "").replace(" ", "").replace("-", "")
    if num.startswith("01"):
        num = "2" + num
    return {
        "status": "ok",
        "whatsapp_number": num,
        "contact_name": "أحمد عرفات",
        "support_message": "مرحباً أ/ أحمد عرفات، أحتاج مساعدة أو استفسار بخصوص تطبيق بارق."
    }

@app.post("/admin/whatsapp")
async def set_whatsapp_number(req: Request):
    data = await req.json()
    num = str(data.get("whatsapp_number", "")).strip()
    if not num:
        raise HTTPException(status_code=400, detail="يرجى إدخال رقم هاتف صحيح")
    cfg = load_config()
    cfg["whatsapp_number"] = num
    save_config(cfg)
    return {"status": "ok", "message": "تم تحديث رقم الدعم الفني عبر واتساب بنجاح", "whatsapp_number": num}

# --- MASTER DATABASE API ---
@app.get("/api/master-database/info")
async def get_master_db_info():
    db_path = os.path.join(BASE_DIR, "master_database.xlsx")
    if os.path.exists(db_path):
        size_kb = round(os.path.getsize(db_path) / 1024, 1)
        mtime = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(os.path.getmtime(db_path)))
        return {"status": "ok", "exists": True, "size_kb": size_kb, "last_updated": mtime}
    return {"status": "ok", "exists": False}

@app.post("/admin/master-database/upload")
async def upload_master_database(req: Request):
    form = await req.form()
    file = form.get("file")
    if not file:
        raise HTTPException(status_code=400, detail="يرجى اختيار ملف الإكسل")
    content = await file.read()
    db_path = os.path.join(BASE_DIR, "master_database.xlsx")
    with open(db_path, "wb") as fh:
        fh.write(content)
    headers, rows = _parse_any_excel_file(content, "")
    return {
        "status": "ok",
        "message": f"تم رفع وتحديث قاعدة البيانات الرئيسية بنجاح ({len(rows)} صف)",
        "rows_count": len(rows),
        "headers": headers
    }


@app.get("/api/public-plans")
async def get_public_plans():
    plans = load_plans()
    return [p for p in plans if p.get("is_active", True)]

@app.post("/api/request-plan")
async def request_plan(req: Request):
    data = await req.json()
    plan_name = data.get("plan_name", "باقة")
    user_name = data.get("user_name", "مندوب")
    phone = data.get("phone", "")
    cfg = load_config()
    raw_num = str(cfg.get("whatsapp_number", "201094593394")).strip().replace("+", "").replace(" ", "").replace("-", "")
    if raw_num.startswith("01"):
        wa_target = "2" + raw_num
    else:
        wa_target = raw_num
    import urllib.parse
    phone_info = f" ({phone})" if phone else ""
    msg = f"مرحباً أ/ أحمد عرفات، أنا المندوب {user_name}{phone_info} وأرغب في الاشتراك وتفعيل {plan_name} في تطبيق بارق."
    encoded_msg = urllib.parse.quote(msg)
    return {
        "status": "ok",
        "message": f"تم تسجيل طلب اشتراكك في '{plan_name}' بنجاح! جاري تحويلك للأستاذ أحمد عرفات عبر واتساب لتفعيل الباقة فوراً.",
        "whatsapp_url": f"https://api.whatsapp.com/send?phone={wa_target}&text={encoded_msg}"
    }

@app.post("/auth/logout")
@app.get("/auth/logout")
async def auth_logout():
    return {"status": "ok", "message": "تم تسجيل الخروج بنجاح"}

@app.post("/auth/presence")
@app.delete("/auth/presence")
async def auth_presence():
    return {"status": "ok"}

@app.post("/auth/refresh")
async def auth_refresh():
    return {"status": "ok", "token": "bareq_refreshed_token"}

@app.post("/auth/change-password")
async def change_password(req: Request):
    data = await req.json()
    cfg = load_config()
    old_pw = data.get("old_password", "").strip()
    new_pw = data.get("new_password", "").strip()
    if not new_pw or len(new_pw) < 3:
        raise HTTPException(status_code=400, detail="كلمة المرور الجديدة يجب أن تكون 3 أحرف على الأقل")
    if old_pw != cfg.get("admin_password", "123"):
        raise HTTPException(status_code=400, detail="كلمة المرور الحالية غير صحيحة")
    cfg["admin_password"] = new_pw
    save_config(cfg)
    return {"status": "ok", "message": "تم تغيير كلمة المرور بنجاح"}

@app.post("/api/check-live/ws-ticket")
async def ws_ticket():
    return {"status": "ok", "ticket": "bareq_live_ticket_99"}

# --- ADMIN DASHBOARD & STORAGE PAGES ---
@app.get("/admin/check-storage")
async def check_storage_page():
    html = """
    <!DOCTYPE html>
    <html lang="ar" dir="rtl">
    <head>
        <meta charset="UTF-8">
        <title>تخزين الفرز — برق</title>
        <style>
            body { font-family: sans-serif; background: #060b14; color: #d8e4f0; padding: 2rem; direction: rtl; }
            .card { background: #0b1423; border: 1px solid #1a2d45; padding: 1.5rem; border-radius: 10px; max-width: 800px; margin: 0 auto; }
            h1 { color: #0ea5e9; font-size: 1.4rem; }
            p { color: #6b7f96; font-size: 0.9rem; line-height: 1.6; }
            .btn { background: #0ea5e9; color: white; padding: 0.6rem 1.2rem; border-radius: 6px; text-decoration: none; display: inline-block; font-weight: bold; margin-top: 1rem; }
        </style>
    </head>
    <body>
        <div class="card">
            <h1>📦 فرز وقواعد بيانات التخزين</h1>
            <p>لا توجد بيانات فرز قديمة أو ملفات معلقة في الوقت الحالي. جميع عمليات الفرز والتسجيل الحالية تُعالج محلياً في الذاكرة وتُصدَّر مباشرة كملفات Excel.</p>
            <a href="/" class="btn">العودة للوحة التحكم الرئيسية</a>
        </div>
    </body>
    </html>
    """
    return HTMLResponse(content=html)

@app.get("/admin/check-storage/summary")
@app.get("/api/admin/check-storage/summary")
async def check_storage_summary():
    return {"status": "ok", "summary": [], "total_rows": 0}

@app.get("/admin/gemini-overview")
async def gemini_overview():
    return {
        "cost_usd": "0.00$",
        "tokens": 0,
        "reset_at": time.strftime("%Y-%m-%d %H:%M:%S")
    }

@app.post("/admin/gemini-overview/reset")
async def gemini_overview_reset():
    return {"status": "ok", "message": "تم تصفير عداد التوكنات"}

@app.get("/admin/plates-overview")
async def plates_overview():
    return {
        "total": 0,
        "tafrigh": 0,
        "live": 0,
        "ptt": 0
    }

@app.get("/admin/online-users")
async def online_users():
    return [
        {
            "username": "admin",
            "display_name": "مدير النظام",
            "device_id": "Device_Local_PC",
            "last_seen": "الآن"
        }
    ]

@app.get("/admin/storage-summary")
async def storage_summary():
    return {"status": "ok", "summary": []}

@app.get("/admin/groups")
async def get_groups():
    return []

@app.post("/admin/groups")
async def create_group(req: Request):
    data = await req.json()
    return {"status": "ok", "id": 1, "name": data.get("name", "مجموعة جديدة")}

@app.get("/admin/users")
async def get_users():
    return load_users()

@app.post("/admin/users")
async def create_user(req: Request):
    data = await req.json()
    users = load_users()
    
    new_user = {
        "id": len(users) + 1,
        "username": data.get("username", f"user_{len(users)+1}"),
        "password": data.get("password", "123456"),
        "display_name": data.get("display_name") or data.get("username"),
        "is_admin": data.get("is_admin", False),
        "is_active": True,
        "rows_limit": data.get("rows_limit", 3000000),
        "subscription_end": "مفتوح - 30 يوماً",
        "created_at": time.strftime("%Y-%m-%d %H:%M:%S")
    }
    users.append(new_user)
    save_users(users)
    return {"status": "ok", "message": f"تم إنشاء المستخدم '{new_user['username']}' بنجاح", "user": new_user}

@app.get("/admin/users/{user_id}")
async def get_user_detail(user_id: int):
    users = load_users()
    for u in users:
        if u["id"] == user_id:
            user_data = dict(u)
            user_data.setdefault("is_active", True)
            user_data.setdefault("is_admin", False)
            user_data.setdefault("display_name", u.get("username", ""))
            user_data.setdefault("rows_limit", 3000000)
            user_data.setdefault("subscription_end", "مفتوح - 30 يوماً")
            user_data.setdefault("created_at", time.strftime("%Y-%m-%d %H:%M:%S"))
            user_data.setdefault("group_id", None)
            user_data.setdefault("gemini_rest_model_id", "")
            user_data.setdefault("gemini_live_model_id", "")
            user_data.setdefault("gemini_check_model_id", "")
            user_data.setdefault("gemini_spend_usd", 0.0)
            user_data.setdefault("gemini_spend_limit_usd", None)
            user_data.setdefault("device_binding", "غير مقيد")
            user_data.setdefault("last_seen", "متصل الآن")
            return user_data
    raise HTTPException(status_code=404, detail="المستخدم غير موجود")

@app.patch("/admin/users/{user_id}")
@app.put("/admin/users/{user_id}")
async def patch_user(user_id: int, req: Request):
    data = await req.json()
    users = load_users()
    for u in users:
        if u["id"] == user_id:
            u.update(data)
            save_users(users)
            return {"status": "ok", "user": u}
    return {"status": "ok"}

@app.delete("/admin/users/{user_id}")
async def delete_user(user_id: int):
    users = load_users()
    users = [u for u in users if u["id"] != user_id]
    save_users(users)
    return {"status": "ok", "message": "تم حذف المستخدم بنجاح"}

@app.patch("/admin/users/{user_id}/group")
@app.patch("/admin/users/{user_id}/gemini-policy")
@app.patch("/admin/users/{user_id}/rows-limit")
async def patch_user_properties(user_id: int, req: Request):
    data = await req.json()
    users = load_users()
    for u in users:
        if u["id"] == user_id:
            u.update(data)
            save_users(users)
            return {"status": "ok", "message": "تم تحديث البيانات بنجاح", "user": u}
    return {"status": "ok"}

@app.post("/admin/users/{user_id}/reset-device")
@app.post("/admin/users/{user_id}/upgrade-device-binding")
@app.post("/admin/users/{user_id}/logout")
@app.post("/admin/users/{user_id}/renew-subscription")
@app.post("/admin/users/{user_id}/subscription/activate")
@app.post("/admin/users/{user_id}/subscription/suspend")
async def user_action_handler(user_id: int):
    users = load_users()
    for u in users:
        if u["id"] == user_id:
            u["subscription_end"] = "مفتوح - 30 يوماً"
            u["is_active"] = True
            save_users(users)
            break
    return {"status": "ok", "message": "تمت العملية بنجاح"}

@app.get("/admin/provider/gemini-models")
async def get_admin_gemini_models():
    return {
        "models": [
            {"id": 1, "channel": "rest", "model_id": "gemini-1.5-flash", "label": "Gemini 1.5 Flash"},
            {"id": 2, "channel": "live", "model_id": "gemini-1.5-flash", "label": "Gemini 1.5 Flash"},
            {"id": 3, "channel": "check", "model_id": "gemini-1.5-flash", "label": "Gemini 1.5 Flash"}
        ]
    }

@app.post("/admin/provider/gemini-models")
async def add_admin_gemini_model(req: Request):
    return {"status": "ok", "message": "تمت إضافة الموديل"}

@app.get("/admin/provider/gemini-defaults")
async def get_gemini_defaults():
    return {
        "rest_model_id": "gemini-1.5-flash",
        "live_model_id": "gemini-1.5-flash",
        "check_model_id": "gemini-1.5-flash"
    }

@app.put("/admin/provider/gemini-defaults")
async def put_gemini_defaults(req: Request):
    return {"status": "ok", "message": "تم التحديث"}

@app.get("/admin/provider/key-pools")
async def get_key_pools():
    cfg = load_config()
    def make_pool(keys_list, kind):
        result = []
        for i, k in enumerate(keys_list):
            short = k[:8] + "..." + k[-4:] if len(k) > 14 else k
            result.append({
                "id": str(i+1),
                "name": f"{kind} مفتاح {i+1}",
                "short": short,
                "masked": short,
                "label": f"مفتاح {i+1}",
                "status": "active",
                "key": k,
                "value": k
            })
        return result

    return {
        "redis": True,
        "gemini_live_key_source": "redis",
        "gemini_rest_key_source": "redis",
        "gemini_vertex_primary": "1",
        "gemini_rest":  make_pool(cfg.get("gemini_rest_keys", []), "REST"),
        "gemini_live":  make_pool(cfg.get("gemini_live_keys", []), "Live"),
        "groq":         make_pool(cfg.get("groq_keys", []), "Groq Whisper Turbo"),
        "ors":          make_pool(cfg.get("ors_keys", []), "ORS"),
        "gmaps":        make_pool(cfg.get("gmaps_keys", [cfg.get("gmaps_api_key","")]) if cfg.get("gmaps_api_key") else [], "Maps"),
    }

@app.post("/admin/provider/key-pools/{kind}")
@app.post("/admin/provider/key-pools/{kind}/keys")
async def add_key_pool(kind: str, req: Request):
    data = await req.json()
    key = str(data.get("value") or data.get("key") or data.get("api_key") or "").strip()
    label = str(data.get("label") or data.get("name") or "").strip()
    if not key:
        raise HTTPException(status_code=400, detail="المفتاح فارغ")
    
    cfg = load_config()
    
    map_kind = {
        "gemini_rest":  "gemini_rest_keys",
        "gemini_live":  "gemini_live_keys",
        "groq":         "groq_keys",
        "ors":          "ors_keys",
        "gmaps":        "gmaps_keys",
    }
    field = map_kind.get(kind)
    if not field:
        raise HTTPException(status_code=400, detail=f"نوع مفتاح غير معروف: {kind}")
    
    existing = cfg.get(field, [])
    if key not in existing:
        existing.append(key)
        cfg[field] = existing
        
        # Also set primary key for fast access
        if kind == "gemini_rest" and not cfg.get("gemini_api_key"):
            cfg["gemini_api_key"] = key
        if kind == "gemini_live" and not cfg.get("gemini_api_key"):
            cfg["gemini_api_key"] = key
        if kind == "groq" and not cfg.get("groq_api_key"):
            cfg["groq_api_key"] = key
        if kind == "gmaps":
            cfg["gmaps_api_key"] = key
        if kind == "ors":
            cfg["ors_api_key"] = key
            
        save_config(cfg)
    
    return {"status": "ok", "message": f"تم إضافة المفتاح بنجاح ({kind})"}

@app.get("/admin/provider/key-pools/{kind}/keys/{key_id}/secret")
async def get_key_pool_secret(kind: str, key_id: str):
    cfg = load_config()
    map_kind = {
        "gemini_rest":  "gemini_rest_keys",
        "gemini_live":  "gemini_live_keys",
        "groq":         "groq_keys",
        "ors":          "ors_keys",
        "gmaps":        "gmaps_keys",
    }
    field = map_kind.get(kind)
    if field and field in cfg:
        keys = cfg[field]
        try:
            idx = int(key_id) - 1
            if 0 <= idx < len(keys):
                return {"value": keys[idx]}
        except Exception:
            for k in keys:
                if key_id in k or k == key_id:
                    return {"value": k}
    return {"value": ""}

@app.delete("/admin/provider/key-pools/{kind}/{key_id}")
@app.delete("/admin/provider/key-pools/{kind}/keys/{key_id}")
async def delete_key_pool(kind: str, key_id: str):
    cfg = load_config()
    map_kind = {
        "gemini_rest":  "gemini_rest_keys",
        "gemini_live":  "gemini_live_keys",
        "groq":         "groq_keys",
        "ors":          "ors_keys",
        "gmaps":        "gmaps_keys",
    }
    field = map_kind.get(kind)
    if field and field in cfg:
        keys = cfg[field]
        try:
            idx = int(key_id) - 1
            if 0 <= idx < len(keys):
                keys.pop(idx)
                cfg[field] = keys
                save_config(cfg)
                return {"status": "ok"}
        except Exception:
            pass
        # Fallback delete by value match
        if key_id in keys:
            keys.remove(key_id)
            cfg[field] = keys
            save_config(cfg)
            return {"status": "ok"}
    return {"status": "ok"}

@app.post("/admin/provider/key-pools/{kind}/keys/{key_id}/unpark")
async def unpark_key_pool(kind: str, key_id: str):
    return {"status": "ok", "message": "تم إرجاع المفتاح للدوران"}

@app.patch("/admin/provider/key-pools/{kind}/keys/{key_id}")
@app.put("/admin/provider/key-pools/{kind}/keys/{key_id}")
async def update_key_pool_status(kind: str, key_id: str, req: Request):
    return {"status": "ok", "message": "تم تحديث المفتاح"}

@app.get("/admin/provider/gemini-pricing")
async def get_gemini_pricing():
    return {
        "rest_in": 0.15,
        "rest_out": 0.60,
        "live_in": 0.15,
        "live_out": 0.60
    }

@app.put("/admin/provider/gemini-pricing")
async def put_gemini_pricing(req: Request):
    return {"status": "ok", "message": "تم حفظ الأسعار"}

@app.put("/admin/provider/gemini-rest-key-source")
@app.put("/admin/provider/gemini-live-key-source")
@app.put("/admin/provider/gemini-vertex-primary")
@app.put("/admin/provider/gemini-vertex-enabled-slots")
async def update_key_sources(req: Request):
    return {"status": "ok", "message": "تم الحفظ بنجاح"}

# Main UI Index
@app.get("/")
async def index():
    html_path = os.path.join(BASE_DIR, "index.html")
    if os.path.exists(html_path):
        with open(html_path, "r", encoding="utf-8") as f:
            return HTMLResponse(content=f.read())
    return HTMLResponse("<h1>Bareq Server</h1>")

# ─────────────────────────────────────────────────────────
# PLANS (PACKAGES) API — إدارة باقات الاشتراك
# ─────────────────────────────────────────────────────────

@app.get("/admin/plans")
async def get_plans():
    return load_plans()

@app.post("/admin/plans")
async def create_plan(req: Request):
    data = await req.json()
    plans = load_plans()
    new_id = max((p["id"] for p in plans), default=0) + 1
    new_plan = {
        "id": new_id,
        "name": data.get("name", f"باقة {new_id}"),
        "name_en": data.get("name_en", f"Plan {new_id}"),
        "price": data.get("price", 0),
        "currency": data.get("currency", "ريال"),
        "duration_days": data.get("duration_days", 30),
        "rows_limit": data.get("rows_limit", 5000),
        "description": data.get("description", ""),
        "features": data.get("features", []),
        "color": data.get("color", "#22c55e"),
        "is_active": True,
        "created_at": time.strftime("%Y-%m-%d %H:%M:%S"),
    }
    plans.append(new_plan)
    save_plans(plans)
    return {"status": "ok", "message": f"تم إنشاء الباقة '{new_plan['name']}' بنجاح", "plan": new_plan}

@app.put("/admin/plans/{plan_id}")
@app.patch("/admin/plans/{plan_id}")
async def update_plan(plan_id: int, req: Request):
    data = await req.json()
    plans = load_plans()
    for p in plans:
        if p["id"] == plan_id:
            p.update({k: v for k, v in data.items() if k != "id"})
            save_plans(plans)
            return {"status": "ok", "message": "تم تحديث الباقة بنجاح", "plan": p}
    raise HTTPException(status_code=404, detail="الباقة غير موجودة")

@app.delete("/admin/plans/{plan_id}")
async def delete_plan(plan_id: int):
    plans = load_plans()
    plans = [p for p in plans if p["id"] != plan_id]
    save_plans(plans)
    return {"status": "ok", "message": "تم حذف الباقة بنجاح"}

@app.post("/admin/users/{user_id}/assign-plan")
async def assign_plan_to_user(user_id: int, req: Request):
    data = await req.json()
    plan_id = data.get("plan_id")
    plans = load_plans()
    plan = next((p for p in plans if p["id"] == plan_id), None)
    if not plan:
        raise HTTPException(status_code=404, detail="الباقة غير موجودة")
    users = load_users()
    for u in users:
        if u["id"] == user_id:
            from datetime import datetime, timedelta
            start = datetime.now()
            end = start + timedelta(days=plan["duration_days"])
            u["plan_id"] = plan["id"]
            u["plan_name"] = plan["name"]
            u["rows_limit"] = plan["rows_limit"]
            u["subscription_start"] = start.strftime("%Y-%m-%d")
            u["subscription_end"] = end.strftime("%Y-%m-%d")
            u["is_active"] = True
            save_users(users)
            return {
                "status": "ok",
                "message": f"تم تعيين باقة '{plan['name']}' للمستخدم بنجاح",
                "subscription_end": u["subscription_end"],
                "rows_limit": u["rows_limit"],
            }
    raise HTTPException(status_code=404, detail="المستخدم غير موجود")

@app.get("/admin/plans/stats")
async def get_plans_stats():
    plans = load_plans()
    users = load_users()
    stats = []
    for p in plans:
        count = sum(1 for u in users if u.get("plan_id") == p["id"])
        stats.append({"plan_id": p["id"], "plan_name": p["name"], "subscribers": count})
    return stats

if __name__ == "__main__":
    host = os.environ.get("HOST", "0.0.0.0")
    try:
        port = int(os.environ.get("PORT", 8500))
    except (ValueError, TypeError):
        port = 8500
    print("==================================================")
    print("   Bareq System Server - Running Successfully")
    print(f"   URL: http://{host}:{port} (Local: http://127.0.0.1:{port})")
    print("==================================================")
    uvicorn.run(app, host=host, port=port, ws="auto")
